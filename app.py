"""
Camp Sol Taplin - Enrollment Dashboard
Flask Application with User Management and Live CampMinder API Integration
"""

from dotenv import load_dotenv
load_dotenv()

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pandas as pd
import os
import json
import traceback
import uuid
from datetime import datetime, date, timedelta
from io import BytesIO
import threading

# Import our custom modules
from parser import CampMinderParser
from historical_data import HistoricalDataManager
from budget_data import BUDGET_FY2026, parse_po_file, build_budget_vs_actual

# Try to import CampMinder API client (optional)
try:
    from campminder_api import CampMinderAPIClient, EnrollmentDataProcessor, FinancialDataProcessor
    CAMPMINDER_API_AVAILABLE = True
except ImportError:
    CAMPMINDER_API_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'camp-sol-taplin-2026-secret-key')

# Configuration
UPLOAD_FOLDER = 'static/uploads'
DATA_FOLDER = 'data'
CACHE_FILE = os.path.join(DATA_FOLDER, 'api_cache.json')
ALLOWED_EXTENSIONS = {'csv'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max
app.config['TEMPLATES_AUTO_RELOAD'] = True

# ==================== DATABASE CONFIG ====================
DB_POOL_SIZE = 5
DB_MAX_OVERFLOW = 10
DB_POOL_RECYCLE_SECONDS = 300

from flask_sqlalchemy import SQLAlchemy

database_url = os.environ.get('DATABASE_URL', 'sqlite:///local.db')
# Render PostgreSQL uses "postgres://" but SQLAlchemy needs "postgresql://"
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# Connection pool settings to handle stale/dropped PostgreSQL connections
if database_url.startswith('postgresql://'):
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,        # Test connections before use (fixes SSL EOF errors)
        'pool_recycle': DB_POOL_RECYCLE_SECONDS,
        'pool_size': DB_POOL_SIZE,
        'max_overflow': DB_MAX_OVERFLOW,
    }

db = SQLAlchemy(app)

# CampMinder API Configuration (from environment variables)
CAMPMINDER_API_KEY = os.environ.get('CAMPMINDER_API_KEY')
CAMPMINDER_SUBSCRIPTION_KEY = os.environ.get('CAMPMINDER_SUBSCRIPTION_KEY')
CAMPMINDER_SEASON_ID = int(os.environ.get('CAMPMINDER_SEASON_ID', '2026'))
CACHE_TTL_MINUTES = 15  # Cache data for 15 minutes
ATTENDANCE_LOCK_HOUR = 17  # 5:00 PM

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# ==================== PERMISSION DEFINITIONS ====================

ALL_PERMISSIONS = [
    'view_dashboard', 'view_bydate', 'view_comparison',
    'view_campcomparison', 'view_detailed', 'view_finance',
    'edit_groups', 'download_excel', 'upload_csv',
    'manage_users', 'manage_settings',
    'take_attendance', 'view_attendance',
    'view_fieldtrips', 'manage_fieldtrips',
]

PERMISSION_LABELS = {
    'view_dashboard': 'Dashboard (Executive Summary)',
    'view_bydate': 'By Date View',
    'view_comparison': 'Year Comparison',
    'view_campcomparison': 'Camp Comparison',
    'view_detailed': 'Detailed View',
    'view_finance': 'Finance Tab',
    'edit_groups': 'Edit Groups',
    'download_excel': 'Download Excel',
    'upload_csv': 'Upload CSV',
    'manage_users': 'Manage Users',
    'manage_settings': 'Settings',
    'take_attendance': 'Take Attendance',
    'view_attendance': 'View Attendance (Admin)',
    'view_fieldtrips': 'View Field Trips',
    'manage_fieldtrips': 'Manage Field Trips',
}

ROLE_DEFAULT_PERMISSIONS = {
    'admin': list(ALL_PERMISSIONS),
    'viewer': [
        'view_dashboard', 'view_bydate', 'view_comparison',
        'view_campcomparison', 'view_detailed', 'download_excel',
        'view_fieldtrips',
    ],
    'unit_leader': [
        'view_campcomparison', 'view_detailed', 'edit_groups',
        'take_attendance',
    ],
}

# ==================== DATABASE MODELS ====================

class UserAccount(db.Model):
    __tablename__ = 'users'
    username = db.Column(db.String(120), primary_key=True)
    password_hash = db.Column(db.String(256), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='viewer')
    permissions = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def get_permissions(self):
        """Return list of permissions from DB, or role defaults if not set."""
        if self.permissions:
            try:
                return json.loads(self.permissions)
            except (json.JSONDecodeError, TypeError):
                pass
        return list(ROLE_DEFAULT_PERMISSIONS.get(self.role, []))

    def has_permission(self, perm):
        """Check if user has a specific permission."""
        return perm in self.get_permissions()

    def set_permissions(self, perms_list):
        self.permissions = json.dumps(perms_list)

class GroupAssignment(db.Model):
    __tablename__ = 'group_assignments'
    id = db.Column(db.Integer, primary_key=True)
    program = db.Column(db.String(100), nullable=False)
    week = db.Column(db.Integer, nullable=False)
    person_id = db.Column(db.String(20), nullable=False)
    group_number = db.Column(db.Integer, nullable=False)
    __table_args__ = (db.UniqueConstraint('program', 'week', 'person_id'),)

class ProgramSetting(db.Model):
    __tablename__ = 'program_settings'
    program = db.Column(db.String(120), primary_key=True)
    goal = db.Column(db.Integer, nullable=False, default=0)
    weeks_offered = db.Column(db.Integer, nullable=False, default=9)
    weeks_active = db.Column(db.String(50), nullable=False, default='1,2,3,4,5,6,7,8,9')
    active = db.Column(db.Boolean, nullable=False, default=True)

class GlobalSetting(db.Model):
    __tablename__ = 'global_settings'
    key = db.Column(db.String(50), primary_key=True)
    value = db.Column(db.Text, nullable=False)

# ==================== FIELD TRIP MODELS ====================

class FieldTripVenue(db.Model):
    __tablename__ = 'field_trip_venues'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, unique=True)
    address = db.Column(db.String(300), nullable=True)
    waiver_url = db.Column(db.String(500), nullable=True)
    active = db.Column(db.Boolean, default=True)

class FieldTripAssignment(db.Model):
    __tablename__ = 'field_trip_assignments'
    id = db.Column(db.Integer, primary_key=True)
    group_name = db.Column(db.String(100), nullable=False)
    week = db.Column(db.Integer, nullable=False)
    venue_id = db.Column(db.Integer, db.ForeignKey('field_trip_venues.id'), nullable=True)
    trip_date = db.Column(db.Date, nullable=True)
    confirmed = db.Column(db.Boolean, default=False)
    comments = db.Column(db.Text, nullable=True)
    buses_ja = db.Column(db.Integer, default=0)
    buses_jcc = db.Column(db.Integer, default=0)
    __table_args__ = (db.UniqueConstraint('group_name', 'week'),)

# ==================== ATTENDANCE MODELS ====================

class UnitLeaderAssignment(db.Model):
    __tablename__ = 'unit_leader_assignments'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(120), db.ForeignKey('users.username'), nullable=False)
    program_name = db.Column(db.String(100), nullable=False)
    __table_args__ = (db.UniqueConstraint('username', 'program_name'),)

class AttendanceCheckpoint(db.Model):
    __tablename__ = 'attendance_checkpoints'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False, unique=True)
    sort_order = db.Column(db.Integer, nullable=False, default=0)
    time_label = db.Column(db.String(20), nullable=True)
    active = db.Column(db.Boolean, nullable=False, default=True)

class AttendanceRecord(db.Model):
    __tablename__ = 'attendance_records'
    id = db.Column(db.Integer, primary_key=True)
    person_id = db.Column(db.String(20), nullable=False)
    program_name = db.Column(db.String(100), nullable=False)
    week = db.Column(db.Integer, nullable=False)
    date = db.Column(db.Date, nullable=False)
    checkpoint_id = db.Column(db.Integer, db.ForeignKey('attendance_checkpoints.id'), nullable=False)
    status = db.Column(db.String(20), nullable=False, default='present')
    recorded_by = db.Column(db.String(120), nullable=False)
    recorded_at = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.String(500), nullable=True)
    __table_args__ = (
        db.UniqueConstraint('person_id', 'program_name', 'date', 'checkpoint_id'),
        db.Index('idx_attendance_date_program', 'date', 'program_name'),
    )

# ==================== INIT DB & DEFAULT USERS ====================

with app.app_context():
    db.create_all()
    # Ensure 'permissions' column exists (for existing databases)
    from sqlalchemy import inspect as sa_inspect
    inspector = sa_inspect(db.engine)
    columns = [c['name'] for c in inspector.get_columns('users')]
    if 'permissions' not in columns:
        with db.engine.connect() as conn:
            conn.execute(db.text("ALTER TABLE users ADD COLUMN permissions TEXT"))
            conn.commit()
    # Migrate global_settings.value from VARCHAR(200) to TEXT (PostgreSQL only)
    if db.engine.dialect.name == 'postgresql':
        gs_columns = {c['name']: c for c in inspector.get_columns('global_settings')}
        if 'value' in gs_columns:
            col_type = str(gs_columns['value'].get('type', ''))
            if 'VARCHAR' in col_type.upper() or 'CHAR' in col_type.upper():
                with db.engine.connect() as conn:
                    conn.execute(db.text("ALTER TABLE global_settings ALTER COLUMN value TYPE TEXT"))
                    conn.commit()
    # Migrate existing users: backfill permissions from role if not set
    for u in UserAccount.query.all():
        if u.permissions is None:
            default_perms = ROLE_DEFAULT_PERMISSIONS.get(u.role, [])
            u.permissions = json.dumps(default_perms)
    db.session.commit()
    # Ensure admin users always have ALL current permissions (catches newly added ones)
    for u in UserAccount.query.filter_by(role='admin').all():
        current_perms = u.get_permissions()
        if set(current_perms) != set(ALL_PERMISSIONS):
            u.permissions = json.dumps(list(ALL_PERMISSIONS))
    db.session.commit()
    # Create default users if they don't exist
    if not UserAccount.query.filter_by(username='campsoltaplin@marjcc.org').first():
        db.session.add(UserAccount(
            username='campsoltaplin@marjcc.org',
            password_hash=generate_password_hash('M@rjcc2026'),
            role='admin'
        ))
    if not UserAccount.query.filter_by(username='onlyview').first():
        db.session.add(UserAccount(
            username='onlyview',
            password_hash=generate_password_hash('M@rjcc2026'),
            role='viewer'
        ))
    # Seed program settings from hardcoded defaults if table is empty
    if ProgramSetting.query.count() == 0:
        DEFAULT_GOALS = {
            'Infants': 6, 'Toddler': 12, 'PK2': 26, 'PK3': 36, 'PK4': 40,
            'Tsofim': 100, "Children's Trust Tsofim": 10,
            'Yeladim': 100, "Children's Trust Yeladim": 10,
            'Chaverim': 75, "Children's Trust Chaverim": 10,
            'Giborim': 60, "Children's Trust Giborim": 10,
            'Madli-Teen': 40, "Children's Trust Madli-Teen": 5,
            'Teen Travel': 30, 'Teen Travel: Epic Trip to Orlando': 15,
            'Basketball': 25, 'Flag Football': 20, 'Soccer': 25,
            'Sports Academy 1': 20, 'Sports Academy 2': 20,
            'Tennis Academy': 20, 'Tennis Academy - Half Day': 15,
            'Swim Academy': 20,
            'Tiny Tumblers Gymnastics': 15, 'Recreational Gymnastics': 20,
            'Competitive Gymnastics Team': 15, 'Volleyball': 20, 'MMA Camp': 15,
            'Teeny Tiny Tnuah': 20, 'Tiny Tnuah 1': 25, 'Tiny Tnuah 2': 25,
            'Tnuah 1': 30, 'Tnuah 2': 30, 'Extreme Tnuah': 20,
            'Art Exploration': 20, 'Music Camp': 20, 'Theater Camp': 25,
            'Madatzim 9th Grade': 25, 'Madatzim 10th Grade': 20,
            'OMETZ': 15
        }
        for prog, goal in DEFAULT_GOALS.items():
            db.session.add(ProgramSetting(program=prog, goal=goal, weeks_offered=9, weeks_active='1,2,3,4,5,6,7,8,9', active=True))
    # Migrate existing ProgramSettings: set weeks_active if missing/empty
    for ps in ProgramSetting.query.all():
        if not ps.weeks_active:
            # Reconstruct weeks_active from weeks_offered (e.g. 7 → "1,2,3,4,5,6,7")
            n = ps.weeks_offered if ps.weeks_offered and 1 <= ps.weeks_offered <= 9 else 9
            ps.weeks_active = ','.join(str(i) for i in range(1, n + 1))
    # Seed global settings if empty
    if not GlobalSetting.query.filter_by(key='total_goal').first():
        db.session.add(GlobalSetting(key='total_goal', value='750'))
    if not GlobalSetting.query.filter_by(key='revenue_goal').first():
        db.session.add(GlobalSetting(key='revenue_goal', value='0'))
    # Seed default attendance checkpoints
    if AttendanceCheckpoint.query.count() == 0:
        db.session.add(AttendanceCheckpoint(name='Morning', sort_order=1, time_label='9:00 AM', active=True))
        db.session.add(AttendanceCheckpoint(name='After Lunch', sort_order=2, time_label='1:00 PM', active=True))
        db.session.add(AttendanceCheckpoint(name='Departure', sort_order=3, time_label='3:30 PM', active=True))
        db.session.add(AttendanceCheckpoint(name='KC Before', sort_order=4, time_label='7:30 AM', active=True))
        db.session.add(AttendanceCheckpoint(name='KC After', sort_order=5, time_label='4:00 PM', active=True))
    else:
        # Ensure KC checkpoints exist (for existing databases)
        if not AttendanceCheckpoint.query.filter_by(name='KC Before').first():
            db.session.add(AttendanceCheckpoint(name='KC Before', sort_order=4, time_label='7:30 AM', active=True))
        if not AttendanceCheckpoint.query.filter_by(name='KC After').first():
            db.session.add(AttendanceCheckpoint(name='KC After', sort_order=5, time_label='4:00 PM', active=True))
    # Ensure Early Pickup checkpoint exists (stores EP flag independently of main status)
    if not AttendanceCheckpoint.query.filter_by(name='Early Pickup').first():
        db.session.add(AttendanceCheckpoint(name='Early Pickup', sort_order=6, time_label='', active=True))

    # ---- Unit Leader accounts ----
    UNIT_LEADER_PERMISSIONS = ['view_dashboard', 'view_detailed', 'edit_groups',
                               'take_attendance', 'view_attendance']
    UNIT_LEADERS = {
        'ecacamp@marjcc.org': [
            'Infants', 'Toddler', 'PK2', 'PK3', 'PK4',
        ],
        'tsofim@marjcc.org': [
            'Tsofim', "Children's Trust Tsofim",
        ],
        'yeladim@marjcc.org': [
            'Yeladim', "Children's Trust Yeladim",
        ],
        'chaverim@marjcc.org': [
            'Chaverim', "Children's Trust Chaverim",
        ],
        'giborim@marjcc.org': [
            'Giborim', "Children's Trust Giborim",
        ],
        'madli-teen@marjcc.org': [
            'Madli-Teen', "Children's Trust Madli-Teen",
        ],
        'teentravel@marjcc.org': [
            'Teen Travel', 'Teen Travel: Epic Trip to Orlando',
        ],
        'basketballcamp@marjcc.org': [
            'Basketball',
        ],
        'flagfootballcamp@marjcc.org': [
            'Flag Football',
        ],
        'soccercamp@marjcc.org': [
            'Soccer',
        ],
        'sportsacademy@marjcc.org': [
            'Sports Academy 1', 'Sports Academy 2',
        ],
        'tennisacademy@marjcc.org': [
            'Tennis Academy', 'Tennis Academy - Half Day',
        ],
        'swimacademy@marjcc.org': [
            'Swim Academy',
        ],
        'gymnastics@marjcc.org': [
            'Tiny Tumblers Gymnastics', 'Recreational Gymnastics',
            'Competitive Gymnastics Team',
        ],
        'volleyballcamp@marjcc.org': [
            'Volleyball',
        ],
        'karatecamp@marjcc.org': [
            'MMA Camp',
        ],
        'tnuah@marjcc.org': [
            'Teeny Tiny Tnuah', 'Tiny Tnuah 1', 'Tiny Tnuah 2',
            'Tnuah 1', 'Tnuah 2', 'Extreme Tnuah',
        ],
        'artexploration@marjcc.org': [
            'Art Exploration',
        ],
        'musicacademy@marjcc.org': [
            'Music Camp',
        ],
        'theatercamp@marjcc.org': [
            'Theater Camp',
        ],
        'madatzim@marjcc.org': [
            'Madatzim 9th Grade', 'Madatzim 10th Grade',
        ],
        'ometz@marjcc.org': [
            'OMETZ',
        ],
    }
    for ul_username, ul_programs in UNIT_LEADERS.items():
        if not UserAccount.query.filter_by(username=ul_username).first():
            u = UserAccount(
                username=ul_username,
                password_hash=generate_password_hash('M@rjcc2026'),
                role='unit_leader',
            )
            u.set_permissions(UNIT_LEADER_PERMISSIONS)
            db.session.add(u)
            db.session.flush()  # ensure user exists before adding assignments
            for prog in ul_programs:
                db.session.add(UnitLeaderAssignment(
                    username=ul_username, program_name=prog))
            print(f"Created unit leader: {ul_username} -> {', '.join(ul_programs)}")

    # ---- Seed Field Trip Venues ----
    if FieldTripVenue.query.count() == 0:
        SEED_VENUES = [
            ('Altitude Coral Springs', '2035 N University Dr, Coral Springs, FL 33071', None),
            ('Anhinga Clay Studios', '4600 SW 75 Avenue, Miami, FL 33155', None),
            ('Arcade Zone Fun Park', '10064 W Oakland Park Blvd, Sunrise, FL 33351', None),
            ('Beach', '1870 S Ocean Dr, Hallandale Beach, FL 33009', None),
            ('Bowlero Dania Point', '1841 Way Pointe Pl, Dania Beach, FL 33004', None),
            ('Boat', '', None),
            ('CB Smith Water Park', '900 N Flamingo Rd, Pembroke Pines, FL 33028', None),
            ('Casa', '', None),
            ('Chuck E. Cheese', '8515 Pines Blvd, Pembroke Pines, FL 33024', None),
            ('Color Me Mine', '13680 W State Rd 84, Davie, FL 33325', None),
            ('Dave and Busters', '3000 Oakwood Blvd, Hollywood, FL 33020', None),
            ('Dezerland Action Park', '14401 NE 19th Ave, North Miami, FL 33181', 'Dezerland Action Park - Waiver.pdf'),
            ('Diver Mansion', '12885 Biscayne Blvd #3, North Miami, FL 33181', 'https://divermansion.com/north-miami/waiver/'),
            ('Everglades Holiday Park', '21940 Griffin Rd, Fort Lauderdale, FL 33332', None),
            ('Flamingo Park', '999 11th St, Miami Beach, FL 33139', None),
            ('Flippos in Ft. Lauderdale', '1455 SE 17th St, Fort Lauderdale, FL 33316', None),
            ('Flying Squirrel Trampoline Park', '3305 Corporate Ave, Weston, FL 33331', 'https://waiver.roller.app/FlyingSquirrelWeston/home'),
            ('Frost Science Museum', '1101 Biscayne Blvd, Miami, FL 33132', None),
            ('Fundimension', '2129 NW 1st Ct, Miami, FL 33127', None),
            ('Game Time', '5701 Sunset Dr #330, South Miami, FL 33143', None),
            ('Kayak at Oleta', '3400 NE 163rd St, North Miami Beach, FL 33160', None),
            ('Kids Empire', '11401 Pines Blvd #270, Pembroke Pines, FL 33026', None),
            ('Le Chocolatier', '1840 NE 164th St, North Miami Beach, FL 33162', None),
            ('Miami Zoo', '12400 SW 152nd St, Miami, FL 33177', None),
            ("Miami Children's Museum", '980 MacArthur Causeway, Miami, FL 33132', None),
            ('Miami Seaquarium', '4400 Rickenbacker Causeway, Miami, FL 33149', None),
            ('Movies', '128 Sunset Dr, Dania Beach, FL 33004', None),
            ('Museum of Discovery & Science', '401 SW 2nd St, Fort Lauderdale, FL 33312', None),
            ('Monstar Mini Golf', '8358 Pines Blvd, Pembroke Pines, FL 33024', None),
            ('MegaJump', '8901 NW 20th St, Doral, FL 33172', None),
            ('Off the Wall', '4939 Coconut Creek Pkwy, Coconut Creek, FL 33063', 'off the wall gameroom waiver.pdf'),
            ('Paradox Museum Miami', '2301 N Miami Ave, Miami, FL 33127', None),
            ('Perez Art Museum Miami', '1103 Biscayne Blvd, Miami, FL 33132', None),
            ('Pines Ice Arena', '12425 Taft St, Pembroke Pines, FL 33028', None),
            ('Pinstripes - Aventura', '19505 Biscayne Boulevard, Miami, FL 33180', None),
            ('Restor Pass', '', None),
            ('PopStroke', '1314 N Federal Hwy, Delray Beach, FL 33483', None),
            ('Puttshack', '701 S Miami Ave, Miami, FL 33131', None),
            ('Rapids', '6566 N Military Trail, Riviera Beach, FL 33407', None),
            ('REVO Indoor Soccer', '10395 NW 41st St, Doral, FL 33178', None),
            ('Shark Wake Park', '1440 Eshleman Trail, West Palm Beach, FL 33413', None),
            ('Spearz Bowling', '5325 S University Dr, Davie, FL 33328', None),
            ('The Edge Rock Gym', '13972 SW 139th Ct, Miami, FL 33186', 'https://drive.google.com/file/d/1y6EquScWMeq2CzKQK-xHRMfH3KvtQ4dl/view'),
            ('The Bass Museum', '2100 Collins Ave, Miami Beach, FL 33139', None),
            ('The Poppet Project', '8650 Biscayne Blvd #29, El Portal, FL 33138', None),
            ('Tidal Cove Waterpark', '19999 W Country Club Dr, Aventura, FL 33180', None),
            ('Tigertail Lake', '580 Gulfstream Way, Dania Beach, FL 33004', None),
            ('Top Golf', '17321 NW 7th Ave, Miami Gardens, FL 33169', None),
            ('Urban Air', '801 South University Drive, Plantation, FL 33324', 'Urban Air.pdf'),
            ("Jumpin' Jamboree", '6000 NW 97th Ave #1, Doral, FL 33178', None),
            ('Jungle Island', '1111 Parrot Jungle Trail, Miami, FL 33132', None),
            ('Volcano Bay', '6000 Universal Blvd, Orlando, FL 32819', None),
            ('Venetian Pool', '2701 De Soto Blvd, Coral Gables, FL 33134', None),
            ('Xtreme Action Park', '5300 Powerline Rd, Fort Lauderdale, FL 33309', 'Xtreme Action Park.pdf'),
            ('You Make Candy', '633 NE 167th Street, North Miami Beach, FL 33162', None),
            ('Young at Art Museum', '751 SW 121st Ave, Davie, FL 33325', None),
        ]
        for vname, vaddr, vwaiver in SEED_VENUES:
            db.session.add(FieldTripVenue(name=vname, address=vaddr, waiver_url=vwaiver, active=True))
        print(f"Seeded {len(SEED_VENUES)} field trip venues")

    # ---- Seed Field Trip Group-Day Mapping ----
    if not GlobalSetting.query.filter_by(key='fieldtrip_group_days').first():
        group_days = {
            'Monday': ['Teen Travel', 'Giborim', 'Madli-Teen'],
            'Tuesday': ['Teen Travel', 'Tnuah', 'Volleyball', 'Tiny Tnuah', 'Tsofim'],
            'Wednesday': ['Teen Travel', 'M&M', 'Tennis', 'Chaverim'],
            'Thursday': ['Teen Travel', 'Gymnastics', 'Art', 'Yeladim', 'Sports Academy', 'Karate'],
            'Friday': ['Teen Travel', 'Soccer', 'Basketball & Flag Football'],
        }
        db.session.add(GlobalSetting(key='fieldtrip_group_days', value=json.dumps(group_days)))
        print("Seeded field trip group-day mapping")

    db.session.commit()

# ==================== CAMP WEEK UTILITIES ====================

# Week date ranges for Camp Sol Taplin 2026
CAMP_WEEK_DATES = {
    1: ('2026-06-08', '2026-06-12'),
    2: ('2026-06-15', '2026-06-19'),
    3: ('2026-06-22', '2026-06-26'),
    4: ('2026-06-29', '2026-07-03'),
    5: ('2026-07-06', '2026-07-10'),
    6: ('2026-07-13', '2026-07-17'),
    7: ('2026-07-20', '2026-07-24'),
    8: ('2026-07-27', '2026-07-31'),
    9: ('2026-08-03', '2026-08-07'),
}

def _parse_date_param(source=None, key='date'):
    """Parse a date string from request args or a dict, defaulting to today."""
    if source is None:
        date_str = request.args.get(key, date.today().isoformat())
    else:
        date_str = source.get(key, date.today().isoformat())
    try:
        return date.fromisoformat(date_str)
    except (ValueError, TypeError):
        return date.today()

def get_current_camp_week(today=None):
    """Return current camp week number (1-9) or None if not during camp."""
    if today is None:
        today = datetime.now().date()
    elif isinstance(today, datetime):
        today = today.date()
    for week_num, (start_str, end_str) in CAMP_WEEK_DATES.items():
        start = date.fromisoformat(start_str)
        end = date.fromisoformat(end_str)
        if start <= today <= end:
            return week_num
    return None

def is_camp_day(today=None):
    """Return True if today is a weekday within a camp week."""
    if today is None:
        today = datetime.now().date()
    elif isinstance(today, datetime):
        today = today.date()
    if today.weekday() >= 5:  # Saturday=5, Sunday=6
        return False
    return get_current_camp_week(today) is not None

class User(UserMixin):
    def __init__(self, username, role, permissions):
        self.id = username
        self.role = role
        self.permissions = permissions  # list of permission strings

    def has_permission(self, perm):
        """Check if user has a specific permission. Admin always has all."""
        if self.role == 'admin':
            return True
        return perm in self.permissions

@login_manager.user_loader
def load_user(username):
    u = db.session.get(UserAccount, username)
    if u:
        return User(u.username, u.role, u.get_permissions())
    return None

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Initialize data managers
parser = CampMinderParser()
historical_manager = HistoricalDataManager()

# Store current report data in memory
current_report = {
    'data': None,
    'generated_at': None,
    'filename': None,
    'source': None  # 'csv' or 'api'
}

# API cache
api_cache = {
    'data': None,
    'fetched_at': None,
    'is_fetching': False
}

# Finance cache (separate, longer TTL)
finance_cache = {
    'data': None,
    'fetched_at': None,
    'is_fetching': False
}
FINANCE_CACHE_TTL_MINUTES = 60

# In-memory persons cache (avoids reading 200KB+ JSON from disk on every click)
_persons_mem_cache = None  # Loaded lazily on first use, then kept in memory

# BAC (Before & After Care) background sync state
_bac_sync_state = {'last_synced_at': None, 'is_syncing': False, 'sync_start': None}
BAC_SYNC_TTL_MINUTES = 60

# PO (Purchase Order) data cache — persisted to data/po_data.json
po_cache = {'data': None, 'uploaded_at': None}
po_cache_path = os.path.join('data', 'po_data.json')
if os.path.exists(po_cache_path):
    try:
        with open(po_cache_path) as f:
            po_cache = json.load(f)
    except Exception:
        po_cache = {'data': None, 'uploaded_at': None}

# ==================== CAMPMINDER API FUNCTIONS ====================

def is_api_configured() -> bool:
    """Check if CampMinder API is configured"""
    return bool(CAMPMINDER_API_KEY and CAMPMINDER_SUBSCRIPTION_KEY and CAMPMINDER_API_AVAILABLE)

def load_api_cache() -> dict:
    """Load cached API data from file"""
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r') as f:
                cache = json.load(f)
                # Check if cache is still valid
                if cache.get('fetched_at'):
                    fetched_at = datetime.fromisoformat(cache['fetched_at'])
                    if datetime.now() - fetched_at < timedelta(minutes=CACHE_TTL_MINUTES):
                        return cache
        except Exception as e:
            print(f"Error loading API cache: {e}")
    return None

def save_api_cache(data: dict):
    """Save API data to cache file"""
    os.makedirs(DATA_FOLDER, exist_ok=True)
    try:
        with open(CACHE_FILE, 'w') as f:
            json.dump(data, f)
    except Exception as e:
        print(f"Error saving API cache: {e}")

def _load_program_settings() -> dict:
    """Load program settings from DB for the enrollment processor"""
    all_settings = ProgramSetting.query.all()
    total_goal_row = GlobalSetting.query.filter_by(key='total_goal').first()
    programs = {}
    for s in all_settings:
        # Compute weeks_offered from weeks_active string
        weeks_active = s.weeks_active or '1,2,3,4,5,6,7,8,9'
        weeks_count = len([w for w in weeks_active.split(',') if w.strip()])
        programs[s.program] = {
            'goal': s.goal,
            'weeks_offered': weeks_count if weeks_count > 0 else 9,
            'weeks_active': weeks_active,
            'active': s.active
        }
    return {
        'programs': programs,
        'total_goal': int(total_goal_row.value) if total_goal_row else 750
    }

def fetch_live_data(force_refresh: bool = False) -> dict:
    """
    Fetch live enrollment data from CampMinder API
    
    Args:
        force_refresh: If True, bypass cache
        
    Returns:
        Processed enrollment data or None
    """
    global api_cache
    
    if not is_api_configured():
        print("CampMinder API not configured")
        return None
    
    # Check cache first (unless force refresh)
    if not force_refresh:
        cached = load_api_cache()
        if cached and cached.get('data'):
            print("Using cached API data")
            api_cache['data'] = cached['data']
            api_cache['fetched_at'] = cached.get('fetched_at')
            if cached.get('retention'):
                api_cache['retention'] = cached['retention']
            # Pre-fetch missing persons in background
            _prefetch_all_persons(cached['data'])
            return cached['data']
    
    # Prevent concurrent fetches
    if api_cache['is_fetching']:
        print("API fetch already in progress")
        return api_cache.get('data')
    
    try:
        api_cache['is_fetching'] = True
        print(f"Fetching live data from CampMinder API (Season {CAMPMINDER_SEASON_ID})...")
        
        # Initialize API client
        client = CampMinderAPIClient(CAMPMINDER_API_KEY, CAMPMINDER_SUBSCRIPTION_KEY)
        
        # Fetch raw data
        raw_data = client.get_enrollment_report(CAMPMINDER_SEASON_ID)
        
        # Load program settings from DB for processing
        db_settings = _load_program_settings()

        # Process into dashboard format
        processor = EnrollmentDataProcessor()
        processed_data = processor.process_enrollment_data(raw_data, program_settings=db_settings)
        
        # Update cache
        fetched_at = datetime.now().isoformat()
        api_cache['data'] = processed_data
        api_cache['fetched_at'] = fetched_at

        print(f"API data fetched successfully: {processed_data['summary']['total_enrollment']} campers")

        # Pre-fetch all person details in background so clicks are instant
        _prefetch_all_persons(processed_data)

        # Calculate retention rate using the same authenticated client
        cache_to_save = {
            'data': processed_data,
            'fetched_at': fetched_at
        }
        try:
            retention_data = client.get_retention_rate(
                current_season=CAMPMINDER_SEASON_ID,
                previous_season=CAMPMINDER_SEASON_ID - 1
            )
            api_cache['retention'] = retention_data
            cache_to_save['retention'] = retention_data
            print(f"Retention rate: {retention_data.get('retention_rate', 0)}%")
        except Exception as ret_err:
            print(f"Warning: Could not calculate retention: {ret_err}")

        # Save to file (includes retention if available)
        save_api_cache(cache_to_save)

        return processed_data

    except Exception as e:
        print(f"Error fetching API data: {e}")
        traceback.print_exc()
        return api_cache.get('data')  # Return cached data if available

    finally:
        api_cache['is_fetching'] = False

def _prefetch_all_persons(enrollment_data):
    """Pre-fetch all person details in a background thread so participant clicks are instant."""
    participants = enrollment_data.get('participants', {})
    if not participants:
        return

    # Collect ALL unique person IDs across all programs/weeks
    all_person_ids = set()
    for program_data in participants.values():
        for week_participants in program_data.values():
            for p in week_participants:
                pid = p.get('person_id')
                if pid:
                    all_person_ids.add(pid)

    if not all_person_ids:
        return

    # Check which ones are missing from cache
    cache = _load_persons_cache()
    missing = [pid for pid in all_person_ids if str(pid) not in cache]

    if not missing:
        print(f"Persons pre-fetch: all {len(all_person_ids)} persons already cached [OK]")
        return

    print(f"Persons pre-fetch: {len(missing)}/{len(all_person_ids)} missing, fetching in background...")

    def _bg_prefetch(app_ctx, pids_to_fetch):
        with app_ctx:
            try:
                cache = _load_persons_cache()
                _fetch_and_cache_persons(pids_to_fetch, cache)
                print(f"Persons pre-fetch complete: {len(pids_to_fetch)} persons fetched [OK]")
            except Exception as e:
                print(f"Persons pre-fetch error: {e}")
                traceback.print_exc()

    t = threading.Thread(
        target=_bg_prefetch,
        args=(app.app_context(), missing),
        daemon=True
    )
    t.start()

def fetch_financial_data(force_refresh: bool = False, enrollment_report: dict = None) -> dict:
    """
    Fetch financial data from CampMinder Financial API.
    Cached separately from enrollment data with 60-min TTL.
    """
    global finance_cache

    if not is_api_configured():
        return None

    # Check cache
    if not force_refresh and finance_cache.get('data') and finance_cache.get('fetched_at'):
        try:
            cached_at = datetime.fromisoformat(finance_cache['fetched_at'])
            if datetime.now() - cached_at < timedelta(minutes=FINANCE_CACHE_TTL_MINUTES):
                print("Using cached finance data")
                return finance_cache['data']
        except Exception:
            pass

    if finance_cache.get('is_fetching'):
        # Safety timeout: reset is_fetching if stuck for more than 3 minutes
        fetch_start = finance_cache.get('fetch_start')
        if fetch_start and (datetime.now() - fetch_start).total_seconds() > 180:
            print("WARNING: is_fetching stuck for >3min, resetting", flush=True)
            finance_cache['is_fetching'] = False
        else:
            return finance_cache.get('data')

    try:
        finance_cache['is_fetching'] = True
        finance_cache['fetch_start'] = datetime.now()
        print(f"Fetching financial data from CampMinder API (Season {CAMPMINDER_SEASON_ID})...", flush=True)

        client = CampMinderAPIClient(CAMPMINDER_API_KEY, CAMPMINDER_SUBSCRIPTION_KEY)

        # Fetch the three financial endpoints
        categories = client.get_financial_categories()
        payment_methods = client.get_payment_methods()
        transactions = client.get_transaction_details(CAMPMINDER_SEASON_ID)

        print(f"Financial data fetched: {len(transactions)} transactions, "
              f"{len(categories)} categories, {len(payment_methods)} payment methods", flush=True)

        # Process financial data
        processor = FinancialDataProcessor()
        finance_data = processor.process_financial_data(
            transactions=transactions,
            categories=categories,
            payment_methods=payment_methods,
            enrollment_report=enrollment_report,
            season=CAMPMINDER_SEASON_ID
        )

        # Also try to fetch historical financial data for comparison
        for hist_season in [2025, 2024]:
            try:
                hist_txns = client.get_transaction_details(hist_season)
                if hist_txns:
                    hist_data = processor.process_financial_data(
                        transactions=hist_txns,
                        categories=categories,
                        payment_methods=payment_methods,
                        enrollment_report=None,
                        season=hist_season
                    )
                    finance_data[f'historical_{hist_season}'] = {
                        'net_revenue': hist_data['summary']['net_revenue'],
                        'gross_revenue': hist_data['summary']['gross_revenue'],
                        'total_discounts': hist_data['summary']['total_discounts'],
                        'timeline': hist_data.get('timeline', [])
                    }
                    print(f"Historical finance {hist_season}: ${hist_data['summary']['net_revenue']:,.2f} net revenue")
            except Exception as e:
                print(f"Could not fetch historical finance for {hist_season}: {e}")

        # Read revenue goal
        try:
            rg = GlobalSetting.query.filter_by(key='revenue_goal').first()
            finance_data['revenue_goal'] = int(rg.value) if rg and rg.value != '0' else 0
        except Exception:
            finance_data['revenue_goal'] = 0

        # Update cache
        finance_cache['data'] = finance_data
        finance_cache['fetched_at'] = datetime.now().isoformat()

        print(f"Finance data processed: ${finance_data['summary']['net_revenue']:,.2f} net revenue", flush=True)
        return finance_data

    except Exception as e:
        print(f"Error fetching financial data: {e}", flush=True)
        traceback.print_exc()
        import sys; sys.stderr.flush()
        return finance_cache.get('data')

    finally:
        finance_cache['is_fetching'] = False


# ==================== ROUTES ====================

@app.route('/')
def index():
    if current_user.is_authenticated:
        if hasattr(current_user, 'role') and current_user.role == 'unit_leader':
            return redirect(url_for('attendance_page'))
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        # Find user (case-insensitive)
        u = UserAccount.query.filter(db.func.lower(UserAccount.username) == username.lower()).first()

        if u and check_password_hash(u.password_hash, password):
            user = User(u.username, u.role, u.get_permissions())
            login_user(user)
            # Unit leaders go directly to attendance page
            if u.role == 'unit_leader':
                return redirect(url_for('attendance_page'))
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ==================== PUBLIC SHARED VIEW ====================

@app.route('/shared/<token>')
def shared_matrix(token):
    """Public read-only enrollment matrix view (no login required)"""
    # Validate token against stored share_token
    stored = GlobalSetting.query.filter_by(key='share_token').first()
    if not stored or stored.value != token:
        return "Not Found", 404

    # Fetch enrollment data (same logic as dashboard route)
    report_data = None
    generated_at = None
    data_source = None

    if is_api_configured():
        api_data = fetch_live_data(force_refresh=False)
        if api_data:
            report_data = api_data
            generated_at = api_cache.get('fetched_at', datetime.now().isoformat())
            if generated_at:
                try:
                    dt = datetime.fromisoformat(generated_at)
                    generated_at = dt.strftime('%B %d, %Y at %I:%M %p')
                except (ValueError, TypeError):
                    pass
            data_source = 'api'

    # Fall back to CSV upload data
    if not report_data and current_report.get('data'):
        report_data = current_report.get('data')
        generated_at = current_report.get('generated_at')
        data_source = 'csv'

    if not report_data:
        return "Not Found", 404

    # Get 2025 program-level data for Old View Stats comparison
    today = datetime.now()
    programs_2025 = historical_manager.get_programs_as_of_date(2025, today.month, today.day)
    programs_2025_map = {}
    if isinstance(programs_2025, list):
        for p in programs_2025:
            if isinstance(p, dict):
                programs_2025_map[p.get('program', '')] = p

    return render_template('shared_matrix.html',
                         report=report_data,
                         generated_at=generated_at,
                         data_source=data_source,
                         programs_2025_map=programs_2025_map)

# ==================== DASHBOARD ====================

@app.route('/dashboard')
@login_required
def dashboard():
    # Determine data source: API (live) > CSV upload > None
    report_data = None
    generated_at = None
    data_source = None
    
    # Try to get data from API first (if configured)
    if is_api_configured():
        api_data = fetch_live_data(force_refresh=False)
        if api_data:
            report_data = api_data
            generated_at = api_cache.get('fetched_at', datetime.now().isoformat())
            if generated_at:
                try:
                    dt = datetime.fromisoformat(generated_at)
                    generated_at = dt.strftime('%B %d, %Y at %I:%M %p')
                except (ValueError, TypeError):
                    pass
            data_source = 'api'
    
    # Fall back to CSV upload data
    if not report_data and current_report.get('data'):
        report_data = current_report.get('data')
        generated_at = current_report.get('generated_at')
        data_source = 'csv'
    
    today = datetime.now()
    comparison_2025 = historical_manager.get_enrollment_as_of_date(2025, today.month, today.day)
    comparison_2024 = historical_manager.get_enrollment_as_of_date(2024, today.month, today.day)
    
    # Get pace comparison if we have current data
    pace_comparison = None
    if report_data:
        pace_comparison = historical_manager.get_pace_comparison(report_data)
    
    # Get historical comparison data (pass 2026 daily for milestones)
    current_daily = report_data.get('date_stats', {}).get('daily', []) if report_data else []
    historical_comparison = historical_manager.get_comparison_data(current_daily=current_daily)
    
    # Get daily data for charts
    historical_data_2025 = historical_manager.get_daily_data(2025)
    historical_data_2024 = historical_manager.get_daily_data(2024)
    
    # Get comparison chart data
    comparison_chart_data = historical_manager.get_weekly_comparison_chart_data()

    # Get Children's Trust stats for historical years
    ct_stats_2024 = historical_manager.get_childrens_trust_stats(2024)
    ct_stats_2025 = historical_manager.get_childrens_trust_stats(2025)

    # Get CT daily data for date-filtered display in By Date tab
    ct_daily_2025 = historical_manager.get_ct_daily_data(2025)
    ct_daily_2024 = historical_manager.get_ct_daily_data(2024)

    # Get 2025 program-level data for OLD VIEW STATS comparison
    # Filter to only include enrollments up to the equivalent date last year
    # (e.g. if today is Feb 14 2026, show 2025 data through Feb 14 2025)
    programs_2025 = historical_manager.get_programs_as_of_date(2025, today.month, today.day)
    # Build a dict for quick lookup by program name
    programs_2025_map = {}
    if isinstance(programs_2025, list):
        for p in programs_2025:
            if isinstance(p, dict):
                programs_2025_map[p.get('program', '')] = p

    # Get financial data from cache only (non-blocking).
    # If no cache, the frontend will auto-trigger /api/finance/refresh via AJAX.
    finance_data = None
    if current_user.has_permission('view_finance') and is_api_configured():
        finance_data = finance_cache.get('data')
        # Kick off background fetch if cache is empty or stale
        if not finance_data:
            def _bg_fetch(app_ctx, report):
                with app_ctx:
                    try:
                        fetch_financial_data(enrollment_report=report)
                    except Exception as e:
                        print(f"Background finance fetch error: {e}", flush=True)
                        import traceback; traceback.print_exc()
            t = threading.Thread(target=_bg_fetch, args=(app.app_context(), report_data), daemon=True)
            t.start()

    # Budget + PO data for Finance tab
    budget_context = {
        'budget': BUDGET_FY2026,
        'po': po_cache.get('data'),
        'po_uploaded_at': po_cache.get('uploaded_at'),
    }

    return render_template('dashboard.html',
                         report=report_data,
                         generated_at=generated_at,
                         data_source=data_source,
                         api_configured=is_api_configured(),
                         comparison_2025=comparison_2025,
                         comparison_2024=comparison_2024,
                         pace_comparison=pace_comparison,
                         historical_comparison=historical_comparison,
                         historical_data_2025=historical_data_2025,
                         historical_data_2024=historical_data_2024,
                         comparison_chart_data=comparison_chart_data,
                         ct_stats_2024=ct_stats_2024,
                         ct_stats_2025=ct_stats_2025,
                         ct_daily_2024=ct_daily_2024,
                         ct_daily_2025=ct_daily_2025,
                         programs_2025_map=programs_2025_map,
                         finance_data=finance_data,
                         budget_data=budget_context,
                         camp_week_dates=CAMP_WEEK_DATES,
                         user=current_user)

# ==================== USER MANAGEMENT ROUTES ====================

@app.route('/admin/users')
@login_required
def admin_users():
    """User management page"""
    if not current_user.has_permission('manage_users'):
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    all_users = UserAccount.query.all()
    user_list = []
    for u in all_users:
        perms = u.get_permissions()
        user_list.append({
            'username': u.username,
            'role': u.role,
            'permissions': perms,
            'perm_count': len(perms),
            'created_at': u.created_at.isoformat() if u.created_at else 'Unknown'
        })

    return render_template('admin_users.html', users=user_list, user=current_user,
                         all_permissions=ALL_PERMISSIONS,
                         permission_labels=PERMISSION_LABELS,
                         role_defaults=ROLE_DEFAULT_PERMISSIONS,
                         active_page='manage_users')

@app.route('/api/users', methods=['GET'])
@login_required
def api_get_users():
    """API: Get all users"""
    if not current_user.has_permission('manage_users'):
        return jsonify({'error': 'Unauthorized'}), 403

    all_users = UserAccount.query.all()
    user_list = []
    for u in all_users:
        perms = u.get_permissions()
        user_list.append({
            'username': u.username,
            'role': u.role,
            'permissions': perms,
            'perm_count': len(perms),
            'created_at': u.created_at.isoformat() if u.created_at else 'Unknown'
        })

    return jsonify({'users': user_list})

@app.route('/api/users', methods=['POST'])
@login_required
def api_create_user():
    """API: Create new user"""
    if not current_user.has_permission('manage_users'):
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    username = data.get('username', '').strip().lower()
    password = data.get('password', '')
    role = data.get('role', 'viewer')
    permissions = data.get('permissions', None)

    # Validation
    if not username or len(username) < 3:
        return jsonify({'error': 'Username must be at least 3 characters'}), 400

    if not password or len(password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400

    if role not in ['admin', 'viewer', 'unit_leader']:
        return jsonify({'error': 'Invalid role'}), 400

    # Check if username is alphanumeric (allow @ for emails)
    clean = username.replace('_', '').replace('.', '').replace('@', '')
    if not clean.isalnum():
        return jsonify({'error': 'Username can only contain letters, numbers, underscores, dots and @'}), 400

    existing = UserAccount.query.filter_by(username=username).first()
    if existing:
        return jsonify({'error': 'Username already exists'}), 400

    # Determine permissions
    if permissions is not None:
        valid_perms = [p for p in permissions if p in ALL_PERMISSIONS]
        perms_json = json.dumps(valid_perms)
    else:
        perms_json = json.dumps(ROLE_DEFAULT_PERMISSIONS.get(role, []))

    # Create user
    new_user = UserAccount(
        username=username,
        password_hash=generate_password_hash(password),
        role=role,
        permissions=perms_json
    )
    db.session.add(new_user)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'User "{username}" created successfully',
        'user': {
            'username': username,
            'role': role
        }
    })

@app.route('/api/users/<username>', methods=['DELETE'])
@login_required
def api_delete_user(username):
    """API: Delete user"""
    if not current_user.has_permission('manage_users'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    username = username.lower()
    
    # Cannot delete yourself
    if username == current_user.id:
        return jsonify({'error': 'Cannot delete your own account'}), 400
    
    u = UserAccount.query.filter_by(username=username).first()
    if not u:
        return jsonify({'error': 'User not found'}), 404

    db.session.delete(u)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'User "{username}" deleted successfully'
    })

@app.route('/api/users/<username>', methods=['PUT'])
@login_required
def api_update_user(username):
    """API: Unified user update — username, password, role, permissions."""
    if not current_user.has_permission('manage_users'):
        return jsonify({'error': 'Unauthorized'}), 403

    username = username.lower()
    u = UserAccount.query.filter_by(username=username).first()
    if not u:
        return jsonify({'error': 'User not found'}), 404

    data = request.get_json()
    new_username = data.get('username', '').strip().lower()
    new_password = data.get('password', '').strip()
    new_role = data.get('role', '').strip()
    permissions = data.get('permissions', None)

    # Validate role if provided
    if new_role and new_role not in ('admin', 'viewer', 'unit_leader'):
        return jsonify({'error': 'Invalid role'}), 400

    # Validate new username if changed
    if new_username and new_username != username:
        if len(new_username) < 3:
            return jsonify({'error': 'Username must be at least 3 characters'}), 400
        clean = new_username.replace('_', '').replace('.', '').replace('@', '')
        if not clean.isalnum():
            return jsonify({'error': 'Username can only contain letters, numbers, underscores, dots and @'}), 400
        if UserAccount.query.filter_by(username=new_username).first():
            return jsonify({'error': 'Username already exists'}), 400

    # Validate password if provided
    if new_password and len(new_password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400

    # Apply changes
    if new_role and new_role != u.role:
        u.role = new_role

    if permissions is not None:
        valid_perms = [p for p in permissions if p in ALL_PERMISSIONS]
        u.set_permissions(valid_perms)

    if new_password:
        u.password_hash = generate_password_hash(new_password)

    # Handle username rename last (PK change)
    if new_username and new_username != username:
        # Update foreign keys in UnitLeaderAssignment
        UnitLeaderAssignment.query.filter_by(username=username).update({'username': new_username})
        # Update attendance records
        AttendanceRecord.query.filter_by(username=username).update({'username': new_username})
        u.username = new_username

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'User updated successfully',
        'username': new_username or username
    })


@app.route('/api/users/<username>/password', methods=['PUT'])
@login_required
def api_change_password(username):
    """API: Change user password"""
    if not current_user.has_permission('manage_users'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    username = username.lower()
    data = request.get_json()
    new_password = data.get('password', '')
    
    if not new_password or len(new_password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400
    
    u = UserAccount.query.filter_by(username=username).first()
    if not u:
        return jsonify({'error': 'User not found'}), 404

    u.password_hash = generate_password_hash(new_password)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Password for "{username}" changed successfully'
    })

@app.route('/api/users/<username>/role', methods=['PUT'])
@login_required
def api_change_role(username):
    """API: Change user role and reset permissions to role defaults"""
    if not current_user.has_permission('manage_users'):
        return jsonify({'error': 'Unauthorized'}), 403

    username = username.lower()

    # Cannot change your own role
    if username == current_user.id:
        return jsonify({'error': 'Cannot change your own role'}), 400

    data = request.get_json()
    new_role = data.get('role', '')

    if new_role not in ['admin', 'viewer', 'unit_leader']:
        return jsonify({'error': 'Invalid role'}), 400

    u = UserAccount.query.filter_by(username=username).first()
    if not u:
        return jsonify({'error': 'User not found'}), 404

    u.role = new_role
    u.set_permissions(ROLE_DEFAULT_PERMISSIONS.get(new_role, []))
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Role for "{username}" changed to {new_role}'
    })

@app.route('/api/users/<username>/permissions', methods=['GET'])
@login_required
def api_get_user_permissions(username):
    """API: Get permissions for a user"""
    if not current_user.has_permission('manage_users'):
        return jsonify({'error': 'Unauthorized'}), 403
    u = UserAccount.query.filter_by(username=username.lower()).first()
    if not u:
        return jsonify({'error': 'User not found'}), 404
    return jsonify({
        'username': u.username,
        'role': u.role,
        'permissions': u.get_permissions()
    })

@app.route('/api/users/<username>/permissions', methods=['PUT'])
@login_required
def api_update_user_permissions(username):
    """API: Update permissions for a user"""
    if not current_user.has_permission('manage_users'):
        return jsonify({'error': 'Unauthorized'}), 403
    username = username.lower()
    data = request.get_json()
    permissions = data.get('permissions', [])
    valid = [p for p in permissions if p in ALL_PERMISSIONS]
    u = UserAccount.query.filter_by(username=username).first()
    if not u:
        return jsonify({'error': 'User not found'}), 404
    u.set_permissions(valid)
    db.session.commit()
    return jsonify({'success': True, 'message': f'Permissions updated for "{username}"'})

# ==================== PROGRAM SETTINGS ROUTES ====================

SETTINGS_ORDER = [
    'Infants', 'Toddler', 'PK2', 'PK3', 'PK4',
    'Tsofim', "Children's Trust Tsofim",
    'Yeladim', "Children's Trust Yeladim",
    'Chaverim', "Children's Trust Chaverim",
    'Giborim', "Children's Trust Giborim",
    'Madli-Teen', "Children's Trust Madli-Teen",
    'Teen Travel', 'Teen Travel: Epic Trip to Orlando',
    'Basketball', 'Flag Football', 'Soccer',
    'Sports Academy 1', 'Sports Academy 2',
    'Tennis Academy', 'Tennis Academy - Half Day',
    'Swim Academy',
    'Tiny Tumblers Gymnastics', 'Recreational Gymnastics',
    'Competitive Gymnastics Team', 'Volleyball', 'MMA Camp',
    'Teeny Tiny Tnuah', 'Tiny Tnuah 1', 'Tiny Tnuah 2',
    'Tnuah 1', 'Tnuah 2', 'Extreme Tnuah',
    'Art Exploration', 'Music Camp', 'Theater Camp',
    'Madatzim 9th Grade', 'Madatzim 10th Grade',
    'OMETZ'
]

def _sort_settings(settings_list):
    """Sort settings by SETTINGS_ORDER"""
    order_map = {name: i for i, name in enumerate(SETTINGS_ORDER)}
    return sorted(settings_list, key=lambda s: order_map.get(s.program, 999))

@app.route('/admin/settings')
@login_required
def admin_settings():
    """Program settings management page"""
    if not current_user.has_permission('manage_settings'):
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    all_settings = _sort_settings(ProgramSetting.query.all())
    total_goal = GlobalSetting.query.filter_by(key='total_goal').first()
    revenue_goal = GlobalSetting.query.filter_by(key='revenue_goal').first()
    return render_template('admin_settings.html',
                         settings=all_settings,
                         total_goal=int(total_goal.value) if total_goal else 750,
                         revenue_goal=int(revenue_goal.value) if revenue_goal and revenue_goal.value != '0' else 0,
                         user=current_user,
                         active_page='camps_goals')

@app.route('/api/settings', methods=['GET'])
@login_required
def api_get_settings():
    """API: Get all program settings"""
    if not current_user.has_permission('manage_settings'):
        return jsonify({'error': 'Unauthorized'}), 403
    all_settings = _sort_settings(ProgramSetting.query.all())
    total_goal = GlobalSetting.query.filter_by(key='total_goal').first()
    revenue_goal = GlobalSetting.query.filter_by(key='revenue_goal').first()
    return jsonify({
        'programs': [{
            'program': s.program,
            'goal': s.goal,
            'weeks_offered': s.weeks_offered,
            'weeks_active': s.weeks_active or '1,2,3,4,5,6,7,8,9',
            'active': s.active
        } for s in all_settings],
        'total_goal': int(total_goal.value) if total_goal else 750,
        'revenue_goal': int(revenue_goal.value) if revenue_goal and revenue_goal.value != '0' else 0
    })

@app.route('/api/settings', methods=['PUT'])
@login_required
def api_update_settings():
    """API: Bulk update program settings"""
    if not current_user.has_permission('manage_settings'):
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    programs = data.get('programs', [])
    total_goal = data.get('total_goal')

    for p in programs:
        weeks_active = p.get('weeks_active', '1,2,3,4,5,6,7,8,9')
        weeks_count = len([w for w in weeks_active.split(',') if w.strip()])
        setting = ProgramSetting.query.filter_by(program=p['program']).first()
        if setting:
            setting.goal = int(p.get('goal', setting.goal))
            setting.weeks_active = weeks_active
            setting.weeks_offered = weeks_count if weeks_count > 0 else 9
            setting.active = bool(p.get('active', setting.active))
        else:
            db.session.add(ProgramSetting(
                program=p['program'],
                goal=int(p.get('goal', 0)),
                weeks_active=weeks_active,
                weeks_offered=weeks_count if weeks_count > 0 else 9,
                active=bool(p.get('active', True))
            ))

    if total_goal is not None:
        gs = GlobalSetting.query.filter_by(key='total_goal').first()
        if gs:
            gs.value = str(int(total_goal))
        else:
            db.session.add(GlobalSetting(key='total_goal', value=str(int(total_goal))))

    # Save revenue goal if provided
    revenue_goal = data.get('revenue_goal')
    if revenue_goal is not None:
        rg = GlobalSetting.query.filter_by(key='revenue_goal').first()
        if rg:
            rg.value = str(int(revenue_goal))
        else:
            db.session.add(GlobalSetting(key='revenue_goal', value=str(int(revenue_goal))))

    db.session.commit()

    # Clear both memory and file cache so dashboard recalculates with new settings
    api_cache['data'] = None
    api_cache['fetched_at'] = None
    # Also clear finance cache so it picks up new revenue_goal
    finance_cache['data'] = None
    finance_cache['fetched_at'] = None
    # Also clear the file cache
    try:
        if os.path.exists(CACHE_FILE):
            os.remove(CACHE_FILE)
    except Exception:
        pass

    return jsonify({'success': True, 'message': 'Settings saved successfully'})

# ==================== SHARE LINK API ====================

@app.route('/api/share-token', methods=['GET'])
@login_required
def api_get_share_token():
    """Get current share token for public matrix view"""
    if not current_user.has_permission('manage_settings'):
        return jsonify({'error': 'Unauthorized'}), 403
    existing = GlobalSetting.query.filter_by(key='share_token').first()
    if existing:
        return jsonify({
            'token': existing.value,
            'url': url_for('shared_matrix', token=existing.value, _external=True)
        })
    return jsonify({'token': None, 'url': None})

@app.route('/api/share-token/generate', methods=['POST'])
@login_required
def api_generate_share_token():
    """Generate a new share token for public matrix view"""
    if not current_user.has_permission('manage_settings'):
        return jsonify({'error': 'Unauthorized'}), 403
    token = str(uuid.uuid4())
    existing = GlobalSetting.query.filter_by(key='share_token').first()
    if existing:
        existing.value = token
    else:
        db.session.add(GlobalSetting(key='share_token', value=token))
    db.session.commit()
    return jsonify({
        'success': True,
        'token': token,
        'url': url_for('shared_matrix', token=token, _external=True)
    })

@app.route('/api/share-token', methods=['DELETE'])
@login_required
def api_revoke_share_token():
    """Revoke the current share token"""
    if not current_user.has_permission('manage_settings'):
        return jsonify({'error': 'Unauthorized'}), 403
    existing = GlobalSetting.query.filter_by(key='share_token').first()
    if existing:
        db.session.delete(existing)
        db.session.commit()
    return jsonify({'success': True, 'message': 'Share link revoked'})

# ==================== CAMPMINDER API ROUTES ====================

@app.route('/api/refresh', methods=['POST'])
@login_required
def api_refresh_data():
    """Refresh data from CampMinder API"""
    if not is_api_configured():
        return jsonify({
            'success': False,
            'error': 'CampMinder API not configured. Please set CAMPMINDER_API_KEY and CAMPMINDER_SUBSCRIPTION_KEY environment variables.'
        }), 400
    
    try:
        print("Manual refresh requested...")
        data = fetch_live_data(force_refresh=True)
        
        if data:
            return jsonify({
                'success': True,
                'message': 'Data refreshed successfully',
                'summary': {
                    'total_enrollment': data['summary']['total_enrollment'],
                    'total_camper_weeks': data['summary']['total_camper_weeks'],
                    'programs_count': len(data.get('programs', []))
                },
                'fetched_at': api_cache.get('fetched_at')
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to fetch data from API'
            }), 500
            
    except Exception as e:
        print(f"Refresh error: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'An internal error occurred'
        }), 500

@app.route('/api/finance/refresh', methods=['POST'])
@login_required
def api_finance_refresh():
    """Refresh financial data from CampMinder Financial API"""
    if not current_user.has_permission('view_finance'):
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    if not is_api_configured():
        return jsonify({
            'success': False,
            'error': 'CampMinder API not configured.'
        }), 400

    try:
        print("Manual finance refresh requested...")
        # Get current enrollment data for cross-referencing
        enrollment_report = api_cache.get('data')
        finance_data = fetch_financial_data(force_refresh=True, enrollment_report=enrollment_report)

        if finance_data:
            return jsonify({
                'success': True,
                'message': 'Financial data refreshed successfully',
                'finance_data': finance_data,
                'fetched_at': finance_cache.get('fetched_at')
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to fetch financial data from API'
            }), 500

    except Exception as e:
        print(f"Finance refresh error: {e}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'An internal error occurred'
        }), 500

def _load_persons_cache():
    """Load the persons cache — from memory if available, otherwise from disk (once)."""
    global _persons_mem_cache
    if _persons_mem_cache is not None:
        return _persons_mem_cache
    persons_cache_file = os.path.join(DATA_FOLDER, 'persons_cache.json')
    persons_cache = {}
    try:
        if os.path.exists(persons_cache_file):
            with open(persons_cache_file, 'r') as f:
                persons_cache = json.load(f)
            print(f"Persons cache loaded from disk: {len(persons_cache)} entries")
    except Exception:
        pass
    _persons_mem_cache = persons_cache
    return _persons_mem_cache

def _fetch_and_cache_persons(pids_to_fetch, persons_cache):
    """Fetch person details from CampMinder API and update the persons cache.

    Fetches camper data, guardian data (emails, names, phones),
    custom fields (Share Group With), and sibling info.

    Args:
        pids_to_fetch: List of person IDs not yet in cache
        persons_cache: Current cache dict (modified in-place)

    Returns:
        Updated persons_cache dict
    """
    persons_cache_file = os.path.join(DATA_FOLDER, 'persons_cache.json')
    try:
        client = CampMinderAPIClient(CAMPMINDER_API_KEY, CAMPMINDER_SUBSCRIPTION_KEY)
        if not client.authenticate():
            print("Failed to authenticate with CampMinder API for persons lookup")
            raise Exception("Authentication failed")

        # Step 1: Batch fetch all campers (with relatives info)
        print(f"Fetching {len(pids_to_fetch)} persons via batch API...")
        camper_results = client.get_persons_batch(
            pids_to_fetch,
            include_contact_details=True,
            include_relatives=True
        )
        print(f"Got {len(camper_results)} camper results from API")

        # Build a map of camper ID -> camper data
        camper_map = {}
        guardian_ids_needed = set()
        for person in camper_results:
            pid = person.get('ID')
            if pid:
                camper_map[pid] = person
                # Collect guardian IDs we need to fetch for emails
                for rel in person.get('Relatives', []):
                    if rel.get('IsGuardian'):
                        guardian_ids_needed.add(rel['ID'])

        # Step 2: Batch fetch all guardians (with contact details for emails/phones,
        #          and relatives to find ward/sibling relationships)
        guardian_map = {}
        if guardian_ids_needed:
            guardian_ids_list = list(guardian_ids_needed)
            print(f"Fetching {len(guardian_ids_list)} guardians via batch API...")
            guardian_results = client.get_persons_batch(
                guardian_ids_list,
                include_contact_details=True,
                include_relatives=True
            )
            for g in guardian_results:
                gid = g.get('ID')
                if gid:
                    guardian_map[gid] = g
            print(f"Got {len(guardian_map)} guardian results from API")

        # Step 3: Load "Share Group With" from uploaded CSV data
        share_group_map = {}  # pid -> value
        sgw_file = os.path.join(DATA_FOLDER, 'share_group.json')
        try:
            if os.path.exists(sgw_file):
                with open(sgw_file, 'r') as f:
                    share_group_map = json.load(f)
                print(f"Loaded {sum(1 for v in share_group_map.values() if v)} "
                      f"Share Group With entries from file")
            else:
                print("No share_group.json file found (upload CSV from CampMinder)")
        except Exception as e:
            print(f"Error loading share_group.json: {e}")

        # Step 3.5: Build sibling map from guardian wards
        # Each guardian's Relatives with IsWard=True are their children
        # Siblings of camper X = all wards of X's guardians, minus X itself
        ward_ids_all = set()  # all child IDs found across all guardians
        guardian_wards = {}  # guardian_id -> list of ward PIDs
        for gid, g_person in guardian_map.items():
            wards = [r['ID'] for r in g_person.get('Relatives', []) if r.get('IsWard')]
            guardian_wards[gid] = wards
            ward_ids_all.update(wards)

        # Fetch ward info (name + DOB) for all wards
        # ward_info: pid -> {'first_name': ..., 'dob': ...}
        ward_info = {}
        # First populate from camper_map (already fetched campers)
        for wid in ward_ids_all:
            if wid in camper_map:
                n = camper_map[wid].get('Name', {})
                ward_info[wid] = {
                    'first_name': n.get('First', ''),
                    'dob': camper_map[wid].get('DateOfBirth', '')
                }
            elif str(wid) in persons_cache:
                info = persons_cache[str(wid)]
                ward_info[wid] = {
                    'first_name': info.get('first_name', ''),
                    'dob': info.get('date_of_birth', '')
                }

        # Fetch remaining wards that we don't have yet (with camper details for DOB)
        unknown_wards = [wid for wid in ward_ids_all if wid not in ward_info]
        if unknown_wards:
            print(f"Fetching info for {len(unknown_wards)} sibling persons...")
            ward_results = client.get_persons_batch(
                unknown_wards,
                include_contact_details=False,
                include_relatives=False,
                include_camper_details=True
            )
            for wp in ward_results:
                wid = wp.get('ID')
                if wid:
                    n = wp.get('Name', {})
                    ward_info[wid] = {
                        'first_name': n.get('First', ''),
                        'dob': wp.get('DateOfBirth', '')
                    }

        # Step 4: Build person_info for each camper
        for pid in pids_to_fetch:
            camper = camper_map.get(pid)
            if camper:
                # Get grade from CampMinder's CamperDetails
                camper_details = camper.get('CamperDetails', {}) or {}
                grade = camper_details.get('CampGradeName', '') or camper_details.get('SchoolGradeName', '')

                # Get guardians from relatives
                relatives = camper.get('Relatives', [])
                guardians = [r for r in relatives if r.get('IsGuardian')]
                guardians.sort(key=lambda r: (not r.get('IsPrimary', False)))

                # Find siblings: collect all wards from this camper's guardians, exclude self
                sibling_list = []  # [{id, first_name, dob}, ...]
                seen_sibling_ids = set()
                for g in guardians:
                    g_id = g.get('ID')
                    for ward_id in guardian_wards.get(g_id, []):
                        if ward_id != pid and ward_id not in seen_sibling_ids:
                            seen_sibling_ids.add(ward_id)
                            wi = ward_info.get(ward_id, {})
                            if wi.get('first_name'):
                                sibling_list.append({
                                    'id': ward_id,
                                    'first_name': wi['first_name'],
                                    'dob': wi.get('dob', '')
                                })

                # Sort siblings by name for display
                sibling_list.sort(key=lambda s: s['first_name'].lower())

                person_info = {
                    'first_name': camper.get('Name', {}).get('First', ''),
                    'last_name': camper.get('Name', {}).get('Last', ''),
                    'f1p1_email': '', 'f1p1_email2': '',
                    'f1p2_email': '', 'f1p2_email2': '',
                    'guardian1_name': '', 'guardian1_phones': '',
                    'guardian2_name': '', 'guardian2_phones': '',
                    'grade': grade,
                    'date_of_birth': camper.get('DateOfBirth', ''),
                    'guardian_ids': [g.get('ID') for g in guardians if g.get('ID')],
                    'share_group_with': share_group_map.get(str(pid), ''),
                    'gender': camper.get('GenderName', ''),
                    'medical_notes': '',
                    'siblings': [s['first_name'] for s in sibling_list],
                    'sibling_details': sibling_list,
                    'aftercare': '',
                    'carpool': ''
                }

                for g_idx, guardian in enumerate(guardians[:2]):
                    g_id = guardian.get('ID')
                    g_person = guardian_map.get(g_id)
                    if g_person:
                        # Guardian emails
                        emails = g_person.get('ContactDetails', {}).get('Emails', [])
                        prefix = 'f1p1' if g_idx == 0 else 'f1p2'
                        if len(emails) > 0:
                            person_info[f'{prefix}_email'] = emails[0].get('Address', '')
                        if len(emails) > 1:
                            person_info[f'{prefix}_email2'] = emails[1].get('Address', '')

                        # Guardian name
                        g_first = g_person.get('Name', {}).get('First', '')
                        g_last = g_person.get('Name', {}).get('Last', '')
                        guardian_name = f"{g_first} {g_last}".strip()

                        # Guardian phone numbers
                        phones = g_person.get('ContactDetails', {}).get('PhoneNumbers', [])
                        phone_numbers = ', '.join(
                            ph.get('Number', '') for ph in phones if ph.get('Number')
                        )

                        g_prefix = 'guardian1' if g_idx == 0 else 'guardian2'
                        person_info[f'{g_prefix}_name'] = guardian_name
                        person_info[f'{g_prefix}_phones'] = phone_numbers

                persons_cache[str(pid)] = person_info
            else:
                persons_cache[str(pid)] = {
                    'first_name': 'Camper', 'last_name': str(pid),
                    'f1p1_email': '', 'f1p1_email2': '',
                    'f1p2_email': '', 'f1p2_email2': '',
                    'guardian1_name': '', 'guardian1_phones': '',
                    'guardian2_name': '', 'guardian2_phones': '',
                    'grade': '',
                    'date_of_birth': '',
                    'guardian_ids': [],
                    'share_group_with': '',
                    'gender': '',
                    'medical_notes': '',
                    'siblings': [],
                    'sibling_details': [],
                    'aftercare': '',
                    'carpool': ''
                }

        # Save updated cache to disk AND update in-memory cache
        global _persons_mem_cache
        _persons_mem_cache = persons_cache
        try:
            with open(persons_cache_file, 'w') as f:
                json.dump(persons_cache, f)
        except Exception:
            pass
        print(f"Persons cache updated: {len(persons_cache)} entries total")
    except Exception as e:
        print(f"Error fetching persons batch: {e}")
        traceback.print_exc()

    return persons_cache

def _load_and_fetch_persons(person_ids):
    """Load persons cache and auto-fetch any missing IDs from CampMinder API."""
    cache = _load_persons_cache()
    to_fetch = [pid for pid in person_ids if pid and str(pid) not in cache]
    if to_fetch and is_api_configured():
        cache = _fetch_and_cache_persons(to_fetch, cache)
    return cache

def _sync_bac_to_cache(persons_cache=None):
    """Sync Before/After Care weeks from CampMinder financial + ECA data into persons_cache.

    Args:
        persons_cache: Existing cache dict. If None, loads from file.

    Returns:
        Updated persons_cache dict with bac_weeks populated.
    """
    import re
    persons_cache_file = os.path.join(DATA_FOLDER, 'persons_cache.json')

    if persons_cache is None:
        persons_cache = _load_persons_cache()

    if not is_api_configured():
        return persons_cache

    try:
        api = CampMinderAPIClient()

        # BAC from financial transactions (the authoritative source)
        print("BAC sync: fetching financial transactions...")
        transactions = api.get_transaction_details(2026)
        bac_persons = {}
        for t in transactions:
            desc = str(t.get('description', ''))
            if 'before and after' not in desc.lower():
                continue
            if t.get('isReversed', False):
                continue
            pid = t.get('personId')
            week_match = re.search(r'Week\s*(\d+)', desc, re.IGNORECASE)
            if pid and week_match:
                bac_persons.setdefault(pid, set()).add(int(week_match.group(1)))

        # Clear stale bac_weeks from ALL persons first, then set only valid ones
        for str_pid in persons_cache:
            if 'bac_weeks' in persons_cache[str_pid]:
                del persons_cache[str_pid]['bac_weeks']

        # Set bac_weeks only for persons with actual BAC financial transactions
        for pid in bac_persons:
            str_pid = str(pid)
            if str_pid in persons_cache:
                persons_cache[str_pid]['bac_weeks'] = sorted(bac_persons[pid])
            else:
                persons_cache[str_pid] = {'bac_weeks': sorted(bac_persons[pid])}

        # Save to file and update in-memory cache
        global _persons_mem_cache
        _persons_mem_cache = persons_cache
        try:
            with open(persons_cache_file, 'w') as f:
                json.dump(persons_cache, f)
        except Exception:
            pass

        _bac_sync_state['last_synced_at'] = datetime.now()
        print(f"BAC sync complete: {len(bac_persons)} persons with BAC financial transactions")

    except Exception as e:
        print(f"Error syncing BAC data: {e}")
        traceback.print_exc()

    return persons_cache


def _ensure_bac_synced_background():
    """Trigger BAC sync in a background thread if stale. Never blocks the caller."""
    global _bac_sync_state

    # Already synced recently?
    if _bac_sync_state['last_synced_at']:
        elapsed = (datetime.now() - _bac_sync_state['last_synced_at']).total_seconds()
        if elapsed < BAC_SYNC_TTL_MINUTES * 60:
            return  # Still fresh

    # Already syncing? (with 3-min stuck timeout)
    if _bac_sync_state['is_syncing']:
        if _bac_sync_state['sync_start'] and \
           (datetime.now() - _bac_sync_state['sync_start']).total_seconds() > 180:
            _bac_sync_state['is_syncing'] = False
        else:
            return

    if not is_api_configured():
        return

    _bac_sync_state['is_syncing'] = True
    _bac_sync_state['sync_start'] = datetime.now()

    def _bg_bac_sync(app_ctx):
        with app_ctx:
            try:
                _sync_bac_to_cache()
                _bac_sync_state['last_synced_at'] = datetime.now()
                print("Background BAC sync complete [OK]")
            except Exception as e:
                print(f"Background BAC sync error: {e}")
                traceback.print_exc()
            finally:
                _bac_sync_state['is_syncing'] = False

    t = threading.Thread(
        target=_bg_bac_sync,
        args=(app.app_context(),),
        daemon=True
    )
    t.start()


@app.route('/api/participants/<program>/<int:week>')
@login_required
def api_participants(program, week):
    """Fetch participant details (names + guardian emails) for a specific program/week"""
    # Get participants from cached report data
    data = api_cache.get('data')
    if not data or 'participants' not in data:
        return jsonify({'participants': [], 'error': 'No data available'}), 404

    program_data = data['participants'].get(program, {})
    participants = program_data.get(str(week), [])

    if not participants:
        return jsonify({'participants': []})

    # Get unique person_ids that need lookup
    person_ids = [p['person_id'] for p in participants]

    # Load persons cache, auto-fetch missing ones from API
    persons_cache = _load_and_fetch_persons(person_ids)

    # Load group assignments for this program/week from DB
    group_map = _get_group_map(program, week)

    # Build enriched participants list
    enriched = []
    for p in participants:
        pid = str(p['person_id'])
        info = persons_cache.get(pid, {})
        enriched.append({
            'person_id': p['person_id'],
            'first_name': info.get('first_name', 'Camper'),
            'last_name': info.get('last_name', str(p['person_id'])),
            'enrollment_date': p.get('enrollment_date', ''),
            'f1p1_email': info.get('f1p1_email', ''),
            'f1p1_email2': info.get('f1p1_email2', ''),
            'f1p2_email': info.get('f1p2_email', ''),
            'f1p2_email2': info.get('f1p2_email2', ''),
            'share_group_with': info.get('share_group_with', ''),
            'gender': info.get('gender', ''),
            'medical_notes': info.get('medical_notes', ''),
            'siblings': ', '.join(info.get('siblings', [])) if isinstance(info.get('siblings'), list) else str(info.get('siblings', '')),
            'aftercare': info.get('aftercare', ''),
            'carpool': info.get('carpool', ''),
            'group': group_map.get(pid, 0)
        })

    # Sort: unassigned first, then by group number, then by last name
    enriched.sort(key=lambda x: (
        0 if x['group'] == 0 else 1,
        x['group'],
        x['last_name'].lower(),
        x['first_name'].lower()
    ))

    return jsonify({'participants': enriched})

@app.route('/api/group-assignment/<program>/<int:week>', methods=['POST'])
@login_required
def api_save_group_assignment(program, week):
    """Save a group assignment for a single camper, propagating forward to subsequent weeks"""
    if not current_user.has_permission('edit_groups'):
        return jsonify({'error': 'Unauthorized'}), 403

    data = request.get_json()
    person_id = str(data.get('person_id', ''))
    group = data.get('group')
    propagate = data.get('propagate_forward', True)

    if not person_id:
        return jsonify({'error': 'person_id is required'}), 400

    # Determine which weeks to update
    weeks_to_update = [week]

    if propagate:
        cached_data = api_cache.get('data')
        if cached_data and 'participants' in cached_data:
            program_data = cached_data['participants'].get(program, {})
            enrolled_weeks = []
            for week_str, participants_list in program_data.items():
                w = int(week_str)
                for p in participants_list:
                    if str(p['person_id']) == person_id:
                        enrolled_weeks.append(w)
                        break
            # Only propagate to weeks >= current week where camper is enrolled
            weeks_to_update = sorted([w for w in enrolled_weeks if w >= week])
            if not weeks_to_update:
                weeks_to_update = [week]

    # Apply the group assignment to all target weeks
    updated_weeks = []
    for w in weeks_to_update:
        existing = GroupAssignment.query.filter_by(program=program, week=w, person_id=person_id).first()

        if group is None or group == 0:
            if existing:
                db.session.delete(existing)
            updated_weeks.append(w)
        else:
            if existing:
                existing.group_number = int(group)
            else:
                db.session.add(GroupAssignment(
                    program=program, week=w, person_id=person_id, group_number=int(group)
                ))
            updated_weeks.append(w)

    db.session.commit()

    return jsonify({
        'success': True,
        'key': f"{program}_{week}",
        'person_id': person_id,
        'group': group,
        'updated_weeks': updated_weeks
    })

@app.route('/api/reset-groups/<program>/<int:week>', methods=['POST'])
@login_required
def api_reset_groups(program, week):
    """Reset all group assignments for a given program/week (set all to unassigned)"""
    if not current_user.has_permission('edit_groups'):
        return jsonify({'error': 'Unauthorized'}), 403

    deleted = GroupAssignment.query.filter_by(program=program, week=week).delete()
    db.session.commit()

    return jsonify({
        'success': True,
        'deleted_count': deleted,
        'program': program,
        'week': week
    })

@app.route('/api/download-by-groups/<program>/<int:week>')
@login_required
def download_by_groups(program, week):
    """Generate and download Excel file organized by groups with attendance columns"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Get participants data
    data = api_cache.get('data')
    if not data or 'participants' not in data:
        return jsonify({'error': 'No data available'}), 404

    program_data = data['participants'].get(program, {})
    participants = program_data.get(str(week), [])

    # Load persons cache, auto-fetch missing ones from API
    all_pids = [p['person_id'] for p in participants]
    persons_cache = _load_and_fetch_persons(all_pids)

    # Load group assignments from DB
    group_map = _get_group_map(program, week)

    # Build camper list
    campers = []
    for p in participants:
        pid = str(p['person_id'])
        info = persons_cache.get(pid, {})
        campers.append({
            'first_name': info.get('first_name', 'Camper'),
            'last_name': info.get('last_name', ''),
            'gender': info.get('gender', ''),
            'medical_notes': info.get('medical_notes', ''),
            'siblings': ', '.join(info.get('siblings', [])) if isinstance(info.get('siblings'), list) else str(info.get('siblings', '')),
            'aftercare': info.get('aftercare', ''),
            'carpool': info.get('carpool', ''),
            'share_group_with': info.get('share_group_with', ''),
            'group': group_map.get(pid, 0)
        })

    # Separate into groups
    groups = {}
    unassigned = []
    for c in campers:
        g = c['group']
        if g and g > 0:
            groups.setdefault(g, []).append(c)
        else:
            unassigned.append(c)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def create_group_sheet(wb, sheet_title, group_label, camper_list):
        ws = wb.create_sheet(title=sheet_title)
        camper_list.sort(key=lambda c: (c['last_name'].lower(), c['first_name'].lower()))

        ws.merge_cells('A1:M1')
        ws['A1'] = f'{program} - Week {week} | {group_label}'
        ws['A1'].font = title_font

        headers = ['#', 'First Name', 'Last Name', 'Gender',
                   'Medical Notes', 'Siblings', 'Share Group With',
                   'AfterCare', 'Carpool', 'M', 'T', 'W', 'T', 'F']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        for idx, camper in enumerate(camper_list, 1):
            row = idx + 3
            values = [idx, camper['first_name'], camper['last_name'],
                      camper['gender'], camper['medical_notes'],
                      camper['siblings'], camper['share_group_with'],
                      camper['aftercare'], camper['carpool'],
                      '', '', '', '', '']
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        for col_letter in ['J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col_letter].width = 5

        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    for group_num in sorted(groups.keys()):
        create_group_sheet(wb, f'Group {group_num}', f'Group {group_num}', groups[group_num])

    if unassigned:
        create_group_sheet(wb, 'Unassigned', 'Unassigned', unassigned)

    if not wb.sheetnames:
        ws = wb.create_sheet(title='No Groups')
        ws['A1'] = 'No group assignments found. Please assign groups first.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{program}_Week{week}_ByGroups.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

def _generate_enrollment_excel(programs):
    """Generate enrollment Excel workbook for given programs.
    Returns (BytesIO output, filename string).
    Raises ValueError if data is unavailable.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # Get participants data from cache
    data = api_cache.get('data')
    if not data or 'participants' not in data:
        raise ValueError('No data available')

    # Collect participants per week across all selected programs
    week_participants = {}
    for program in programs:
        program_data = data['participants'].get(program, {})
        for week_str, participants in program_data.items():
            week_num = int(week_str)
            if week_num not in week_participants:
                week_participants[week_num] = []
            for p in participants:
                week_participants[week_num].append(p['person_id'])

    if not week_participants:
        raise ValueError('No enrollment data found for selected programs')

    # Load persons cache and fetch any missing persons
    persons_cache = _load_persons_cache()

    all_pids = set()
    for pids in week_participants.values():
        all_pids.update(pids)

    pids_to_fetch = [
        pid for pid in all_pids
        if str(pid) not in persons_cache or 'guardian1_name' not in persons_cache.get(str(pid), {})
    ]

    if pids_to_fetch and is_api_configured():
        persons_cache = _fetch_and_cache_persons(pids_to_fetch, persons_cache)

    # Build per-week program lookup for ALL enrolled campers
    all_participants = data.get('participants', {})
    pid_programs_by_week = {}
    for prog_name, prog_weeks in all_participants.items():
        for week_str, plist in prog_weeks.items():
            wk = int(week_str)
            if wk not in pid_programs_by_week:
                pid_programs_by_week[wk] = {}
            for p in plist:
                pid = p['person_id']
                if pid not in pid_programs_by_week[wk]:
                    pid_programs_by_week[wk][pid] = []
                pid_programs_by_week[wk][pid].append(prog_name)

    # Load Share Group With data (from uploaded CSV)
    share_group_data = {}
    sgw_file = os.path.join(DATA_FOLDER, 'share_group.json')
    if os.path.exists(sgw_file):
        try:
            with open(sgw_file, 'r') as f:
                share_group_data = json.load(f)
        except Exception:
            pass

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for week_num in sorted(week_participants.keys()):
        person_ids = week_participants[week_num]
        if not person_ids:
            continue

        week_prog_map = pid_programs_by_week.get(week_num, {})

        # Deduplicate by person_id
        seen = set()
        campers = []
        for pid in person_ids:
            if pid in seen:
                continue
            seen.add(pid)
            info = persons_cache.get(str(pid), {})

            # --- Siblings: first names only ---
            sibling_details = info.get('sibling_details', [])
            siblings_raw = info.get('siblings', [])
            if isinstance(siblings_raw, list):
                sibling_first_names = list(siblings_raw)
            else:
                sibling_first_names = [s.strip() for s in str(siblings_raw).split(',') if s.strip()]

            # --- Find youngest sibling ---
            camper_dob = info.get('date_of_birth', '')
            family_members = [{'id': pid, 'first_name': info.get('first_name', ''), 'dob': camper_dob}]
            family_members.extend(sibling_details)

            youngest_id = None
            youngest_name = ''
            youngest_dob = ''
            for member in family_members:
                m_dob = member.get('dob', '')
                if m_dob and (not youngest_dob or m_dob > youngest_dob):
                    youngest_dob = m_dob
                    youngest_id = member.get('id')
                    youngest_name = member.get('first_name', '')

            display_siblings = []
            for sname in sibling_first_names:
                if sname == youngest_name and youngest_id != pid:
                    display_siblings.append(f'*{sname}')
                else:
                    display_siblings.append(sname)

            youngest_program = ''
            if youngest_id and youngest_id != pid:
                progs = week_prog_map.get(youngest_id, [])
                youngest_program = ', '.join(progs) if progs else ''

            campers.append({
                'pid': pid,
                'first_name': info.get('first_name', 'Camper'),
                'last_name': info.get('last_name', ''),
                'grade': info.get('grade', ''),
                'f1p1_email': info.get('f1p1_email', ''),
                'f1p1_email2': info.get('f1p1_email2', ''),
                'f1p2_email': info.get('f1p2_email', ''),
                'f1p2_email2': info.get('f1p2_email2', ''),
                'guardian1_name': info.get('guardian1_name', ''),
                'guardian1_phones': info.get('guardian1_phones', ''),
                'guardian2_name': info.get('guardian2_name', ''),
                'guardian2_phones': info.get('guardian2_phones', ''),
                'siblings': ', '.join(display_siblings),
                'youngest_program': youngest_program,
                'share_group_with': (share_group_data.get(str(pid), '')
                                     or info.get('share_group_with', '')),
            })

        campers.sort(key=lambda c: (c['last_name'].lower(), c['first_name'].lower()))

        ws = wb.create_sheet(title=f'W{week_num}')

        program_list = ', '.join(programs)
        ws.merge_cells('A1:O1')
        ws['A1'] = f'Week {week_num} — {program_list}'
        ws['A1'].font = title_font

        ws['A2'] = f'Total: {len(campers)} campers'
        ws['A2'].font = Font(italic=True, size=10, color='666666')

        headers = ['#', 'Last Name', 'First Name', 'Grade',
                   'Email 1', 'Email 2', 'Email 3', 'Email 4',
                   'Guardian 1', 'Guardian 1 Phone',
                   'Guardian 2', 'Guardian 2 Phone',
                   'Siblings', 'Youngest Sibling Program',
                   'Share Group With']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        for idx, camper in enumerate(campers, 1):
            row = idx + 4
            values = [
                idx, camper['last_name'], camper['first_name'], camper['grade'],
                camper['f1p1_email'], camper['f1p1_email2'],
                camper['f1p2_email'], camper['f1p2_email2'],
                camper['guardian1_name'], camper['guardian1_phones'],
                camper['guardian2_name'], camper['guardian2_phones'],
                camper['siblings'], camper['youngest_program'],
                camper['share_group_with'],
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border

        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 26
        ws.column_dimensions['F'].width = 26
        ws.column_dimensions['G'].width = 26
        ws.column_dimensions['H'].width = 26
        ws.column_dimensions['I'].width = 22
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 22
        ws.column_dimensions['L'].width = 18
        ws.column_dimensions['M'].width = 24
        ws.column_dimensions['N'].width = 28
        ws.column_dimensions['O'].width = 30

        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1

    if not wb.sheetnames:
        ws = wb.create_sheet(title='No Data')
        ws['A1'] = 'No enrollment data found for the selected programs.'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_names = '_'.join(p.replace(' ', '')[:15] for p in programs[:3])
    if len(programs) > 3:
        safe_names += f'_and_{len(programs)-3}_more'
    filename = f"Enrollment_{safe_names}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return output, filename


@app.route('/api/download-multi-program-enrollment', methods=['POST'])
@login_required
def download_multi_program_enrollment():
    """Generate and download enrollment Excel for selected programs."""
    if not current_user.has_permission('download_excel'):
        return jsonify({'error': 'Unauthorized'}), 403

    req_data = request.get_json()
    programs = req_data.get('programs', [])
    if not programs:
        return jsonify({'error': 'No programs selected'}), 400

    try:
        output, filename = _generate_enrollment_excel(programs)
    except ValueError as e:
        return jsonify({'error': str(e)}), 404

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/print-by-groups/<program>/<int:week>')
@login_required
def print_by_groups(program, week):
    """Render print-friendly page showing groups with attendance columns"""
    data = api_cache.get('data')
    if not data or 'participants' not in data:
        return "No data available", 404

    program_data = data['participants'].get(program, {})
    participants = program_data.get(str(week), [])

    all_pids = [p['person_id'] for p in participants]
    persons_cache = _load_and_fetch_persons(all_pids)

    group_map = _get_group_map(program, week)

    campers = []
    for p in participants:
        pid = str(p['person_id'])
        info = persons_cache.get(pid, {})
        campers.append({
            'first_name': info.get('first_name', 'Camper'),
            'last_name': info.get('last_name', ''),
            'gender': info.get('gender', ''),
            'medical_notes': info.get('medical_notes', ''),
            'siblings': ', '.join(info.get('siblings', [])) if isinstance(info.get('siblings'), list) else str(info.get('siblings', '')),
            'aftercare': info.get('aftercare', ''),
            'carpool': info.get('carpool', ''),
            'share_group_with': info.get('share_group_with', ''),
            'group': group_map.get(pid, 0)
        })

    groups = {}
    unassigned = []
    for c in campers:
        g = c['group']
        if g and g > 0:
            groups.setdefault(g, []).append(c)
        else:
            unassigned.append(c)

    # Sort each group
    for g in groups:
        groups[g].sort(key=lambda c: (c['last_name'].lower(), c['first_name'].lower()))
    unassigned.sort(key=lambda c: (c['last_name'].lower(), c['first_name'].lower()))

    return render_template('print_by_groups.html',
                         program=program,
                         week=week,
                         groups=groups,
                         unassigned=unassigned)

@app.route('/api/status')
@login_required
def api_status():
    """Get API configuration status"""
    # Show partial keys for debugging (safe to show first/last few chars)
    api_key_preview = None
    sub_key_preview = None
    
    if CAMPMINDER_API_KEY:
        api_key_preview = f"{CAMPMINDER_API_KEY[:10]}...{CAMPMINDER_API_KEY[-5:]}" if len(CAMPMINDER_API_KEY) > 15 else "TOO_SHORT"
    
    if CAMPMINDER_SUBSCRIPTION_KEY:
        sub_key_preview = f"{CAMPMINDER_SUBSCRIPTION_KEY[:8]}...{CAMPMINDER_SUBSCRIPTION_KEY[-4:]}" if len(CAMPMINDER_SUBSCRIPTION_KEY) > 12 else "TOO_SHORT"
    
    return jsonify({
        'api_configured': is_api_configured(),
        'campminder_api_module_available': CAMPMINDER_API_AVAILABLE,
        'api_key_set': bool(CAMPMINDER_API_KEY),
        'api_key_length': len(CAMPMINDER_API_KEY) if CAMPMINDER_API_KEY else 0,
        'api_key_preview': api_key_preview,
        'subscription_key_set': bool(CAMPMINDER_SUBSCRIPTION_KEY),
        'subscription_key_length': len(CAMPMINDER_SUBSCRIPTION_KEY) if CAMPMINDER_SUBSCRIPTION_KEY else 0,
        'subscription_key_preview': sub_key_preview,
        'season_id': CAMPMINDER_SEASON_ID,
        'cache_ttl_minutes': CACHE_TTL_MINUTES,
        'last_fetch': api_cache.get('fetched_at'),
        'is_fetching': api_cache.get('is_fetching', False),
        'has_cached_data': bool(api_cache.get('data'))
    })

@app.route('/api/test-auth')
@login_required
def api_test_auth():
    """Test CampMinder API authentication - for debugging"""
    if not is_api_configured():
        return jsonify({
            'success': False,
            'error': 'API not configured',
            'api_key_set': bool(CAMPMINDER_API_KEY),
            'subscription_key_set': bool(CAMPMINDER_SUBSCRIPTION_KEY)
        })
    
    import requests
    
    url = "https://api.campminder.com/auth/apikey"
    # Note: CampMinder auth endpoint does NOT want 'Bearer ' prefix
    headers = {
        "Authorization": CAMPMINDER_API_KEY,
        "Ocp-Apim-Subscription-Key": CAMPMINDER_SUBSCRIPTION_KEY
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=30)
        
        return jsonify({
            'success': response.status_code == 200,
            'status_code': response.status_code,
            'response_text': response.text[:500] if response.text else None,
            'response_headers': dict(response.headers),
            'request_url': url,
            'api_key_used': f"{CAMPMINDER_API_KEY[:10]}...{CAMPMINDER_API_KEY[-5:]}" if CAMPMINDER_API_KEY else None,
            'subscription_key_used': f"{CAMPMINDER_SUBSCRIPTION_KEY[:8]}..." if CAMPMINDER_SUBSCRIPTION_KEY else None
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'error_type': type(e).__name__
        })

@app.route('/api/debug-data')
@login_required
def api_debug_data():
    """Debug endpoint to see raw API data"""
    # Check if API is configured
    api_key = os.environ.get('CAMPMINDER_API_KEY')
    sub_key = os.environ.get('CAMPMINDER_SUBSCRIPTION_KEY')
    season_id = int(os.environ.get('CAMPMINDER_SEASON_ID', '2026'))
    
    # First check - are env vars set?
    if not api_key or not sub_key:
        return jsonify({
            'error': 'Environment variables not set',
            'api_key_exists': bool(api_key),
            'sub_key_exists': bool(sub_key)
        })
    
    # Second check - is campminder_api module available?
    if not CAMPMINDER_API_AVAILABLE:
        return jsonify({
            'error': 'campminder_api module not available - check if file exists and has no import errors'
        })
    
    try:
        from campminder_api import CampMinderAPIClient
        
        client = CampMinderAPIClient(api_key, sub_key)
        
        # Authenticate first
        auth_result = client.authenticate()
        if not auth_result:
            return jsonify({
                'error': 'Authentication failed',
                'client_id': client.client_id,
                'jwt_token_exists': bool(client.jwt_token)
            })
        
        client_id = client.client_id
        
        # Get sessions
        sessions = client.get_sessions(season_id, client_id)
        
        # Get programs  
        programs = client.get_programs(season_id, client_id)
        
        # Get attendees (status=6 means Enrolled+Applied)
        attendees = client.get_attendees(season_id, client_id, status=6)
        
        # Create a simple session map: ID -> Name
        session_map = {s['ID']: s['Name'] for s in sessions} if sessions else {}
        
        # Create program map: ID -> Name
        program_map = {p['ID']: p['Name'] for p in programs} if programs else {}
        
        # Get unique session IDs from attendees
        attendee_session_ids = set()
        for att in (attendees or []):
            for sps in att.get('SessionProgramStatus', []):
                attendee_session_ids.add(sps.get('SessionID'))
        
        # Check which session IDs from attendees are missing from sessions
        missing_session_ids = attendee_session_ids - set(session_map.keys())
        
        return jsonify({
            'success': True,
            'client_id': client_id,
            'season_id': season_id,
            'sessions_count': len(sessions) if sessions else 0,
            'sessions_all': [{'id': s['ID'], 'name': s['Name'], 'sort': s.get('SortOrder')} for s in (sessions or [])],
            'programs_count': len(programs) if programs else 0,
            'programs_all': [{'id': p['ID'], 'name': p['Name']} for p in (programs or [])],
            'attendees_count': len(attendees) if attendees else 0,
            'attendees_sample': attendees[:3] if attendees else [],
            'unique_session_ids_in_attendees': list(attendee_session_ids),
            'missing_session_ids': list(missing_session_ids),
            'session_map_sample': dict(list(session_map.items())[:10])
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        })

# ==================== UPLOAD & DATA ROUTES ====================

@app.route('/api/upload-po', methods=['POST'])
@login_required
def upload_po():
    """Upload PO Excel file, parse and store spending data for Budget vs Actual."""
    if not current_user.has_permission('view_finance'):
        return jsonify({'error': 'Unauthorized'}), 403

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'No file provided'}), 400

    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Please upload a .xlsx file'}), 400

    try:
        file_bytes = file.read()
        categories = parse_po_file(file_bytes)
        budget_vs_actual = build_budget_vs_actual(categories)

        po_cache['data'] = {
            'categories': categories,
            'budget_vs_actual': budget_vs_actual,
        }
        po_cache['uploaded_at'] = datetime.now().isoformat()

        # Persist to disk
        os.makedirs('data', exist_ok=True)
        with open(po_cache_path, 'w') as f:
            json.dump(po_cache, f, default=str)

        total_spent = budget_vs_actual['totals']['actual']
        return jsonify({
            'success': True,
            'message': f'PO data loaded: {len(categories)} categories, ${total_spent:,.0f} total spent',
            'data': po_cache['data'],
            'uploaded_at': po_cache['uploaded_at']
        })
    except Exception as e:
        import traceback
        print(f"PO upload error: {traceback.format_exc()}")
        return jsonify({'error': f'Failed to parse PO file: {str(e)}'}), 500

@app.route('/api/upload-share-group', methods=['POST'])
@login_required
def upload_share_group():
    """Upload CampMinder 'Share Group With' CSV export.
    Expected columns: PersonID, Full Name, Share Group With
    Saves to data/share_group.json and updates persons_cache.
    """
    import csv
    import io

    if not current_user.has_permission('upload_csv'):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    try:
        # Read CSV content (handle BOM and multiline fields)
        raw = file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(raw))
        header = next(reader)

        # Find column indices
        header_lower = [h.strip().lower() for h in header]
        pid_col = None
        sgw_col = None
        for i, h in enumerate(header_lower):
            if h == 'personid':
                pid_col = i
            elif 'share group' in h:
                sgw_col = i

        if pid_col is None or sgw_col is None:
            return jsonify({
                'success': False,
                'error': 'CSV must have "PersonID" and "Share Group With" columns'
            }), 400

        # Parse rows
        share_group_data = {}
        count = 0
        for row in reader:
            if len(row) <= max(pid_col, sgw_col):
                continue
            pid = row[pid_col].strip().strip('"')
            sgw = row[sgw_col].strip().strip('"')
            if pid and pid.isdigit():
                # Normalize: replace newlines with comma-space
                sgw_clean = ', '.join(
                    line.strip() for line in sgw.split('\n') if line.strip()
                )
                share_group_data[pid] = sgw_clean
                if sgw_clean:
                    count += 1

        # Save to JSON file
        sgw_file = os.path.join(DATA_FOLDER, 'share_group.json')
        with open(sgw_file, 'w') as f:
            json.dump(share_group_data, f)

        # Also update persons_cache if it exists
        persons_cache = _load_persons_cache()
        updated = 0
        for pid_str, sgw_val in share_group_data.items():
            if pid_str in persons_cache:
                persons_cache[pid_str]['share_group_with'] = sgw_val
                updated += 1

        if updated > 0:
            persons_cache_file = os.path.join(DATA_FOLDER, 'persons_cache.json')
            with open(persons_cache_file, 'w') as f:
                json.dump(persons_cache, f)

        print(f"Share Group With: loaded {count} entries with data, "
              f"updated {updated} cached persons")

        return jsonify({
            'success': True,
            'message': f'Imported {count} Share Group With entries. '
                       f'Updated {updated} cached campers.'
        })

    except Exception as e:
        print(f"Error processing Share Group CSV: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'An internal error occurred'}), 500


@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    """Handle CSV file upload and processing"""
    print("=" * 50)
    print("UPLOAD REQUEST RECEIVED")
    print("=" * 50)
    
    try:
        if not current_user.has_permission('upload_csv'):
            print("ERROR: User lacks upload_csv permission")
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403
        
        if 'file' not in request.files:
            print("ERROR: No file in request")
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        print(f"File received: {file.filename}")
        
        if file.filename == '':
            print("ERROR: Empty filename")
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            print(f"ERROR: Invalid file type: {file.filename}")
            return jsonify({'success': False, 'error': 'Invalid file type. Please upload a CSV file.'}), 400
        
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        saved_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], saved_filename)
        
        print(f"Saving file to: {filepath}")
        file.save(filepath)
        print(f"File saved successfully")
        
        if not os.path.exists(filepath):
            print("ERROR: File not saved")
            return jsonify({'success': False, 'error': 'Failed to save file'}), 500
        
        file_size = os.path.getsize(filepath)
        print(f"File size: {file_size} bytes")
        
        if file_size == 0:
            print("ERROR: File is empty")
            return jsonify({'success': False, 'error': 'File is empty'}), 400
        
        print("Parsing CSV...")
        try:
            report_data = parser.parse_csv(filepath)
            print(f"Parse successful!")
            print(f"  - Total enrollment: {report_data['summary']['total_enrollment']}")
            print(f"  - Total camper weeks: {report_data['summary']['total_camper_weeks']}")
            print(f"  - Programs found: {len(report_data['programs'])}")
        except Exception as parse_error:
            print(f"PARSE ERROR: {str(parse_error)}")
            print(traceback.format_exc())
            return jsonify({
                'success': False, 
                'error': f'Error parsing CSV: {str(parse_error)}'
            }), 500
        
        current_report['data'] = report_data
        current_report['generated_at'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        current_report['filename'] = saved_filename
        
        print("Report data stored successfully")
        print("=" * 50)
        
        return jsonify({
            'success': True,
            'message': 'File processed successfully',
            'redirect': url_for('dashboard'),
            'summary': {
                'total_enrollment': report_data['summary']['total_enrollment'],
                'total_camper_weeks': report_data['summary']['total_camper_weeks'],
                'programs_count': len(report_data['programs'])
            }
        })
        
    except Exception as e:
        print(f"UNEXPECTED ERROR: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False, 
            'error': f'Server error: {str(e)}'
        }), 500

@app.route('/api/retention')
@login_required
def api_retention():
    """Return cached retention rate data.

    Retention is calculated during enrollment cache refresh using the same
    authenticated API client to avoid concurrency/404 issues.
    """
    _ensure_enrollment_cache()

    retention = api_cache.get('retention')
    if retention:
        return jsonify(retention)

    return jsonify({'error': 'Retention data not yet available. Try refreshing.'}), 503


@app.route('/api/recent-enrollments')
@login_required
def recent_enrollments():
    """Return the last 10 campers who registered, with program and week info."""
    _ensure_enrollment_cache()
    if not api_cache.get('data') or not api_cache['data'].get('participants'):
        return jsonify({'enrollments': []})

    participants = api_cache['data']['participants']
    persons_cache = _load_persons_cache()

    # Flatten all participants across programs/weeks, collecting per-person info
    person_map = {}  # person_id -> {name, enrollment_date, programs: {prog: [weeks]}}
    for program, weeks_data in participants.items():
        for week_str, camper_list in weeks_data.items():
            for camper in camper_list:
                pid = str(camper.get('person_id') or camper.get('personId', ''))
                if not pid:
                    continue
                enroll_date = camper.get('enrollment_date', '')

                if pid not in person_map:
                    # Get name from persons_cache or from participant data
                    entry = persons_cache.get(pid, {})
                    if isinstance(entry, dict):
                        name = f"{entry.get('first_name', '')} {entry.get('last_name', '')}".strip()
                    else:
                        name = camper.get('first_name', '') + ' ' + camper.get('last_name', '')
                        name = name.strip()
                    if not name:
                        name = f'Camper {pid}'
                    person_map[pid] = {
                        'name': name,
                        'enrollment_date': enroll_date,
                        'programs': {}
                    }
                else:
                    # Keep the latest enrollment_date
                    if enroll_date and enroll_date > person_map[pid]['enrollment_date']:
                        person_map[pid]['enrollment_date'] = enroll_date

                # Add program/week
                if program not in person_map[pid]['programs']:
                    person_map[pid]['programs'][program] = []
                try:
                    wk = int(week_str)
                except ValueError:
                    wk = week_str
                if wk not in person_map[pid]['programs'][program]:
                    person_map[pid]['programs'][program].append(wk)

    # Sort by enrollment_date descending, take top 10
    sorted_persons = sorted(
        person_map.values(),
        key=lambda p: p['enrollment_date'] or '',
        reverse=True
    )[:10]

    # Format output
    results = []
    for p in sorted_persons:
        # Summarize programs
        prog_summaries = []
        total_weeks = 0
        for prog, weeks in p['programs'].items():
            prog_summaries.append(prog)
            total_weeks += len(weeks)
        results.append({
            'name': p['name'],
            'enrollment_date': p['enrollment_date'],
            'programs': ', '.join(prog_summaries),
            'total_weeks': total_weeks
        })

    return jsonify({'enrollments': results})


@app.route('/api/report-data')
@login_required
def get_report_data():
    """API endpoint to get current report data as JSON"""
    if current_report['data'] is None:
        return jsonify({'error': 'No report data available'}), 404
    
    today = datetime.now()
    comparison_2025 = historical_manager.get_enrollment_as_of_date(2025, today.month, today.day)
    comparison_2024 = historical_manager.get_enrollment_as_of_date(2024, today.month, today.day)
    
    return jsonify({
        'report': current_report['data'],
        'generated_at': current_report['generated_at'],
        'comparison_2025': comparison_2025,
        'comparison_2024': comparison_2024
    })

@app.route('/api/program-comparison/<program_name>')
@login_required
def get_program_comparison(program_name):
    """Get historical comparison data for a specific program"""
    # Get 2026 data from API or current report
    data_2026 = None
    if is_api_configured() and api_cache.get('data'):
        data_2026 = api_cache['data']
    elif current_report.get('data'):
        data_2026 = current_report['data']
    
    # Find program in 2026 data
    program_2026 = None
    if data_2026 and data_2026.get('programs'):
        for prog in data_2026['programs']:
            if prog['program'] == program_name:
                program_2026 = prog
                break
    
    # Get 2025 historical data
    program_2025 = historical_manager.get_program_data(2025, program_name)
    
    # Get 2024 historical data  
    program_2024 = historical_manager.get_program_data(2024, program_name)
    
    return jsonify({
        'program_name': program_name,
        'data_2026': program_2026,
        'data_2025': program_2025,
        'data_2024': program_2024
    })

@app.route('/download-excel')
@login_required
def download_excel():
    """Generate and download Excel report"""
    if not current_user.has_permission('download_excel'):
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    if current_report['data'] is None:
        flash('No report data available', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_data = {
                'Metric': ['Total Enrollment', 'Total Camper Weeks', 'Total FTC', 'Goal', '% to Goal'],
                'Value': [
                    current_report['data']['summary']['total_enrollment'],
                    current_report['data']['summary']['total_camper_weeks'],
                    current_report['data']['summary']['total_fte'],
                    current_report['data']['summary']['goal'],
                    f"{current_report['data']['summary']['percent_to_goal']:.1f}%"
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            programs_df = pd.DataFrame(current_report['data']['programs'])
            programs_df.to_excel(writer, sheet_name='Programs', index=False)
            
            categories_df = pd.DataFrame(current_report['data']['categories'])
            categories_df.to_excel(writer, sheet_name='By Category', index=False)
        
        output.seek(0)
        
        filename = f"camp_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

# ==================== ATTENDANCE ROUTES ====================

def _ensure_enrollment_cache():
    """Ensure enrollment data is loaded in api_cache (from file if needed).
    Unlike load_api_cache(), this ignores TTL — we always want enrollment data available.
    """
    global api_cache
    if not api_cache.get('data') or not api_cache['data'].get('participants'):
        # Try load_api_cache first (respects TTL)
        cached = load_api_cache()
        if cached and cached.get('data'):
            api_cache['data'] = cached['data']
            api_cache['fetched_at'] = cached.get('fetched_at')
        else:
            # Force-load from file ignoring TTL (for attendance — always need enrollment data)
            if os.path.exists(CACHE_FILE):
                try:
                    with open(CACHE_FILE, 'r') as f:
                        cached = json.load(f)
                    if cached and cached.get('data'):
                        api_cache['data'] = cached['data']
                        api_cache['fetched_at'] = cached.get('fetched_at')
                except Exception as e:
                    print(f"Error force-loading enrollment cache: {e}")

def _check_program_access(program_name):
    """Return 403 response if user lacks program access, else None. Admins always pass."""
    if program_name == 'Kid Connection':
        return None  # KC is cross-program, accessible to all authenticated users
    if current_user.role != 'admin':
        assignment = UnitLeaderAssignment.query.filter_by(
            username=current_user.id, program_name=program_name
        ).first()
        if not assignment:
            return jsonify({'error': 'Not authorized for this program'}), 403
    return None

def _get_group_map(program, week):
    """Return dict mapping person_id -> group_number for a program/week."""
    ga_rows = GroupAssignment.query.filter_by(program=program, week=week).all()
    return {ga.person_id: ga.group_number for ga in ga_rows}

def _get_active_checkpoints():
    """Return list of active attendance checkpoints sorted by order."""
    return AttendanceCheckpoint.query.filter_by(active=True).order_by(AttendanceCheckpoint.sort_order).all()

@app.route('/attendance')
@login_required
def attendance_page():
    """Standalone mobile-first attendance page for unit leaders."""
    _ensure_enrollment_cache()
    _ensure_bac_synced_background()

    # Preload basic camper data for the user's assigned programs (instant JS render)
    preloaded = {}
    if api_cache.get('data') and api_cache['data'].get('participants'):
        participants = api_cache['data']['participants']

        if current_user.role == 'admin':
            user_programs = list(participants.keys())
        else:
            assignments = UnitLeaderAssignment.query.filter_by(
                username=current_user.id
            ).all()
            user_programs = [a.program_name for a in assignments]

        persons_cache = _load_persons_cache()

        for prog in user_programs:
            if prog not in participants:
                continue
            prog_weeks = {}
            for week_str, week_campers in participants[prog].items():
                camper_basics = []
                for c in week_campers:
                    pid = str(c.get('personId') or c.get('person_id', ''))
                    entry = persons_cache.get(pid, {})
                    fn = entry.get('first_name', '') if isinstance(entry, dict) else ''
                    ln = entry.get('last_name', '') if isinstance(entry, dict) else ''
                    name = f'{fn} {ln}'.strip() or f'Camper {pid}'
                    has_kc = False
                    if isinstance(entry, dict):
                        bac_weeks = entry.get('bac_weeks')
                        if isinstance(bac_weeks, list):
                            try:
                                has_kc = int(week_str) in bac_weeks
                            except ValueError:
                                pass
                    camper_basics.append({
                        'person_id': pid,
                        'name': name,
                        'has_kc': has_kc,
                    })
                prog_weeks[week_str] = camper_basics
            preloaded[prog] = prog_weeks

    return render_template('attendance.html', preloaded_campers=preloaded)

@app.route('/api/attendance/my-programs')
@login_required
def attendance_my_programs():
    """Return programs assigned to the current unit leader."""
    _ensure_enrollment_cache()
    if current_user.role == 'admin':
        # Admin sees all programs from enrollment data
        programs = []
        if api_cache.get('data') and api_cache['data'].get('participants'):
            programs = sorted(api_cache['data']['participants'].keys())
        return jsonify({'programs': programs})
    # Unit leader: return only assigned programs
    assignments = UnitLeaderAssignment.query.filter_by(username=current_user.id).all()
    programs = [a.program_name for a in assignments]
    return jsonify({'programs': sorted(programs)})

@app.route('/api/attendance/checkpoints')
@login_required
def attendance_checkpoints():
    """Return active checkpoints."""
    checkpoints = _get_active_checkpoints()
    return jsonify({'checkpoints': [
        {'id': c.id, 'name': c.name, 'time_label': c.time_label, 'sort_order': c.sort_order}
        for c in checkpoints
    ]})

@app.route('/api/attendance/campers/<program>/<int:week>')
@login_required
def attendance_campers(program, week):
    """Return camper list for a program/week with attendance for a specific date."""
    # Check access: admin can see all, unit_leader only assigned programs
    denied = _check_program_access(program)
    if denied:
        return denied

    _ensure_enrollment_cache()

    # Parse date from query param (default to today)
    target_date = _parse_date_param()

    # Get campers from enrollment data
    campers = []
    if api_cache.get('data') and api_cache['data'].get('participants'):
        participants = api_cache['data']['participants']
        week_str = str(week)
        if program in participants and week_str in participants[program]:
            campers = participants[program][week_str]

    # Load person names from cache, auto-fetch missing ones from API
    all_pids = [c.get('personId') or c.get('person_id') for c in campers]
    persons_map = _load_and_fetch_persons(all_pids)

    # Trigger BAC sync in background if stale (never blocks this response)
    _ensure_bac_synced_background()

    # Get attendance records for specified date + program
    records = AttendanceRecord.query.filter_by(
        program_name=program, date=target_date
    ).all()
    # Build nested lookup: person_id → {checkpoint_id_str → record}  (O(1) per camper)
    attendance_by_person = {}
    for r in records:
        person_att = attendance_by_person.setdefault(r.person_id, {})
        person_att[str(r.checkpoint_id)] = {
            'status': r.status,
            'notes': r.notes,
            'recorded_at': r.recorded_at.isoformat() if r.recorded_at else None
        }

    # Helper to extract name from persons_cache entry (dict with first_name/last_name)
    def _person_name(person_id, fallback=''):
        entry = persons_map.get(str(person_id))
        if isinstance(entry, dict):
            fn = entry.get('first_name', '')
            ln = entry.get('last_name', '')
            return f'{fn} {ln}'.strip() or fallback
        elif isinstance(entry, str):
            return entry or fallback
        return fallback

    # Determine KC (Kid Connection / Before & After Care) eligibility per camper
    # Primary source: 'bac_weeks' list in persons_cache (synced from CampMinder financial API + ECA sessions)
    # Fallback: legacy fields for backward compatibility

    def _has_kc(person_id):
        entry = persons_map.get(str(person_id))
        if not isinstance(entry, dict):
            return False
        # Primary: check bac_weeks list (populated by /api/attendance/sync-bac)
        bac_weeks = entry.get('bac_weeks')
        if isinstance(bac_weeks, list) and week in bac_weeks:
            return True
        return False

    # Build flat reverse index: person_id → program_name for O(1) sibling lookup
    enrolled_person_to_program = {}
    if api_cache.get('data') and api_cache['data'].get('participants'):
        all_participants = api_cache['data']['participants']
        for prog_name, weeks_data in all_participants.items():
            if str(week) in weeks_data:
                for c in weeks_data[str(week)]:
                    pid = str(c.get('personId') or c.get('person_id', ''))
                    if pid:
                        enrolled_person_to_program[pid] = prog_name

    def _find_youngest_enrolled_sibling(person_id):
        """Find the youngest sibling (younger than camper) enrolled this week."""
        entry = persons_map.get(str(person_id))
        if not isinstance(entry, dict):
            return None
        sibling_details = entry.get('sibling_details')
        if not sibling_details or not isinstance(sibling_details, list):
            return None
        camper_dob = entry.get('date_of_birth', '')
        enrolled_siblings = []
        for sib in sibling_details:
            sib_id = str(sib.get('id', ''))
            if not sib_id:
                continue
            sib_dob = sib.get('dob', '')
            if camper_dob and sib_dob and sib_dob <= camper_dob:
                continue
            # O(1) lookup instead of iterating all programs
            sib_program = enrolled_person_to_program.get(sib_id)
            if sib_program:
                sib_name = _person_name(sib_id, sib.get('first_name', ''))
                enrolled_siblings.append({
                    'name': sib_name,
                    'program': sib_program,
                    'dob': sib_dob
                })
        if not enrolled_siblings:
            return None
        enrolled_siblings.sort(key=lambda s: s.get('dob', ''), reverse=True)
        youngest = enrolled_siblings[0]
        return {'name': youngest['name'], 'program': youngest['program']}

    # Load group assignments for sorting by group
    group_map = _get_group_map(program, week)

    # Build camper list with attendance data
    camper_list = []
    for camper in campers:
        person_id = camper.get('personId') or camper.get('person_id', '')
        name = _person_name(person_id, camper.get('name', f'Camper {person_id}'))
        camper_data = {
            'person_id': str(person_id),
            'name': name,
            'has_kc': _has_kc(person_id),
            'group_number': group_map.get(str(person_id), 0),
            'attendance': attendance_by_person.get(str(person_id), {})
        }
        # Youngest enrolled sibling
        sib = _find_youngest_enrolled_sibling(person_id)
        if sib:
            camper_data['youngest_sibling'] = sib
        camper_list.append(camper_data)

    # Sort by last name
    camper_list.sort(key=lambda c: c['name'].split()[-1] if c['name'] else '')

    return jsonify({
        'program': program,
        'week': week,
        'date': target_date.isoformat(),
        'campers': camper_list,
        'total': len(camper_list)
    })

@app.route('/api/attendance/kc')
@login_required
def attendance_kc():
    """Return all KC-eligible campers for a date, split into ECA vs Other Programs."""
    _ensure_enrollment_cache()
    target_date = _parse_date_param()
    week_num = get_current_camp_week(target_date)
    if not week_num:
        return jsonify({'eca': [], 'other': [], 'date': target_date.isoformat(), 'week': None})

    week_str = str(week_num)
    ECA_PROGRAMS = ['Infants', 'Toddler', 'PK2', 'PK3', 'PK4']

    # Load persons cache + trigger BAC sync in background if stale
    persons_map = _load_persons_cache()
    _ensure_bac_synced_background()

    # Gather all enrolled campers across all programs for this week
    kc_campers = []
    seen_pids = set()
    participants = {}
    if api_cache.get('data') and api_cache['data'].get('participants'):
        participants = api_cache['data']['participants']

    for prog_name, weeks_data in participants.items():
        if week_str not in weeks_data:
            continue
        for camper in weeks_data[week_str]:
            pid = str(camper.get('personId') or camper.get('person_id', ''))
            if not pid or pid in seen_pids:
                continue

            # Check KC eligibility
            entry = persons_map.get(pid)
            if not isinstance(entry, dict):
                continue
            bac_weeks = entry.get('bac_weeks')
            if not isinstance(bac_weeks, list) or week_num not in bac_weeks:
                continue

            seen_pids.add(pid)
            fn = entry.get('first_name', '')
            ln = entry.get('last_name', '')
            name = f'{fn} {ln}'.strip() or f'Camper {pid}'

            is_eca = prog_name in ECA_PROGRAMS
            kc_campers.append({
                'person_id': pid,
                'name': name,
                'program': prog_name,
                'is_eca': is_eca
            })

    # Fetch KC attendance records for this date
    kc_records = AttendanceRecord.query.filter(
        AttendanceRecord.date == target_date,
        AttendanceRecord.checkpoint_id.in_([4, 5]),
        AttendanceRecord.program_name == 'Kid Connection'
    ).all()
    kc_att_map = {}
    for r in kc_records:
        kc_att_map.setdefault(r.person_id, {})[str(r.checkpoint_id)] = {
            'status': r.status,
            'recorded_at': r.recorded_at.isoformat() if r.recorded_at else None
        }

    for c in kc_campers:
        c['attendance'] = kc_att_map.get(c['person_id'], {})

    eca_list = sorted([c for c in kc_campers if c['is_eca']],
                      key=lambda c: c['name'].split()[-1].lower() if c['name'] else '')
    other_list = sorted([c for c in kc_campers if not c['is_eca']],
                        key=lambda c: c['name'].split()[-1].lower() if c['name'] else '')

    return jsonify({
        'eca': eca_list,
        'other': other_list,
        'date': target_date.isoformat(),
        'week': week_num
    })


@app.route('/api/attendance/record', methods=['POST'])
@login_required
def attendance_record():
    """Save a single attendance record (debounced from UI)."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    person_id = data.get('person_id')
    program_name = data.get('program_name')
    checkpoint_id = data.get('checkpoint_id')
    status = data.get('status', 'present')
    notes = data.get('notes', '')

    if not all([person_id, program_name, checkpoint_id]):
        return jsonify({'error': 'Missing required fields'}), 400

    if status not in ('present', 'absent', 'late', 'early_pickup', 'unmarked'):
        return jsonify({'error': 'Invalid status'}), 400

    # Check access
    denied = _check_program_access(program_name)
    if denied:
        return denied

    # Parse target date from request (default to today)
    target_date = _parse_date_param(source=data)

    # Server-side 5 PM lock enforcement (admins bypass)
    if current_user.role != 'admin':
        now = datetime.now()
        today = date.today()
        if target_date < today:
            return jsonify({'error': 'Cannot modify attendance for past days'}), 403
        if target_date == today and now.hour >= ATTENDANCE_LOCK_HOUR:
            return jsonify({'error': 'Day is locked after 5:00 PM'}), 403

    # Handle unmark: delete the record entirely
    if status == 'unmarked':
        existing = AttendanceRecord.query.filter_by(
            person_id=str(person_id),
            program_name=program_name,
            date=target_date,
            checkpoint_id=int(checkpoint_id)
        ).first()
        if existing:
            db.session.delete(existing)
            db.session.commit()
        return jsonify({'success': True, 'action': 'deleted'})

    current_week = get_current_camp_week(target_date)
    if current_week is None:
        # Allow recording even outside camp weeks (for testing/flexibility)
        current_week = 0

    # Upsert: update if exists, insert if not
    existing = AttendanceRecord.query.filter_by(
        person_id=str(person_id),
        program_name=program_name,
        date=target_date,
        checkpoint_id=int(checkpoint_id)
    ).first()

    if existing:
        existing.status = status
        existing.notes = notes
        existing.recorded_by = current_user.id
        existing.recorded_at = datetime.utcnow()
    else:
        record = AttendanceRecord(
            person_id=str(person_id),
            program_name=program_name,
            week=current_week,
            date=target_date,
            checkpoint_id=int(checkpoint_id),
            status=status,
            recorded_by=current_user.id,
            notes=notes
        )
        db.session.add(record)

    try:
        db.session.commit()
        return jsonify({'success': True, 'status': status})
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'error': 'An internal error occurred'}), 500

@app.route('/api/attendance/record-batch', methods=['POST'])
@login_required
def attendance_record_batch():
    """Batch save attendance records (Mark All Present)."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    program_name = data.get('program_name')
    checkpoint_id = data.get('checkpoint_id')
    status = data.get('status', 'present')
    person_ids = data.get('person_ids', [])

    if not all([program_name, checkpoint_id, person_ids]):
        return jsonify({'error': 'Missing required fields'}), 400

    # Check access
    denied = _check_program_access(program_name)
    if denied:
        return denied

    # Parse target date from request (default to today)
    target_date = _parse_date_param(source=data)

    # Server-side 5 PM lock enforcement (admins bypass)
    if current_user.role != 'admin':
        now = datetime.now()
        today = date.today()
        if target_date < today:
            return jsonify({'error': 'Cannot modify attendance for past days'}), 403
        if target_date == today and now.hour >= ATTENDANCE_LOCK_HOUR:
            return jsonify({'error': 'Day is locked after 5:00 PM'}), 403

    # Handle unmark: delete all matching records
    if status == 'unmarked':
        deleted = 0
        for pid in person_ids:
            existing = AttendanceRecord.query.filter_by(
                person_id=str(pid),
                program_name=program_name,
                date=target_date,
                checkpoint_id=int(checkpoint_id)
            ).first()
            if existing:
                db.session.delete(existing)
                deleted += 1
        db.session.commit()
        return jsonify({'success': True, 'count': deleted, 'action': 'deleted'})

    current_week = get_current_camp_week(target_date) or 0

    count = 0
    for pid in person_ids:
        existing = AttendanceRecord.query.filter_by(
            person_id=str(pid),
            program_name=program_name,
            date=target_date,
            checkpoint_id=int(checkpoint_id)
        ).first()
        if existing:
            existing.status = status
            existing.recorded_by = current_user.id
            existing.recorded_at = datetime.utcnow()
        else:
            db.session.add(AttendanceRecord(
                person_id=str(pid),
                program_name=program_name,
                week=current_week,
                date=target_date,
                checkpoint_id=int(checkpoint_id),
                status=status,
                recorded_by=current_user.id
            ))
        count += 1

    try:
        db.session.commit()
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({'error': 'An internal error occurred'}), 500

@app.route('/api/attendance/summary')
@login_required
def attendance_summary():
    """Admin: aggregated attendance stats for all programs on a given date."""
    _ensure_enrollment_cache()
    target_date = _parse_date_param()

    # Get all records for the date
    records = AttendanceRecord.query.filter_by(date=target_date).all()
    checkpoints = _get_active_checkpoints()

    # Get all programs from enrollment data
    all_programs = []
    total_campers_by_program = {}
    if api_cache.get('data') and api_cache['data'].get('participants'):
        participants = api_cache['data']['participants']
        # Figure out which week this date falls in
        week_num = get_current_camp_week(target_date)
        if week_num:
            week_str = str(week_num)
            for prog_name, weeks in participants.items():
                if week_str in weeks and len(weeks[week_str]) > 0:
                    all_programs.append(prog_name)
                    total_campers_by_program[prog_name] = len(weeks[week_str])

    # Build summary: program × checkpoint → {present, absent, late, early_pickup, total}
    summary = {}
    totals = {'present': 0, 'absent': 0, 'late': 0, 'early_pickup': 0}
    for r in records:
        key = (r.program_name, r.checkpoint_id)
        if key not in summary:
            summary[key] = {'present': 0, 'absent': 0, 'late': 0, 'early_pickup': 0, 'marked': 0}
        summary[key][r.status] = summary[key].get(r.status, 0) + 1
        summary[key]['marked'] += 1
        # Only count checkpoint 1 (daily attendance) in KPI totals to avoid double-counting KC
        if r.checkpoint_id == 1:
            totals[r.status] = totals.get(r.status, 0) + 1

    # Format response
    programs_data = []
    for prog in sorted(all_programs):
        total_campers = total_campers_by_program.get(prog, 0)
        prog_entry = {
            'program': prog,
            'total_campers': total_campers,
            'checkpoints': []
        }
        for cp in checkpoints:
            key = (prog, cp.id)
            stats = summary.get(key, {'present': 0, 'absent': 0, 'late': 0, 'early_pickup': 0, 'marked': 0})
            stats['checkpoint_id'] = cp.id
            stats['checkpoint_name'] = cp.name
            stats['total'] = total_campers
            stats['completion'] = round(stats['marked'] / total_campers * 100) if total_campers > 0 else 0
            prog_entry['checkpoints'].append(stats)
        programs_data.append(prog_entry)

    return jsonify({
        'date': target_date.isoformat(),
        'week': get_current_camp_week(target_date),
        'totals': totals,
        'total_campers': sum(total_campers_by_program.values()),
        'programs': programs_data,
        'checkpoints': [{'id': c.id, 'name': c.name, 'time_label': c.time_label} for c in checkpoints]
    })

@app.route('/api/attendance/trends')
@login_required
def attendance_trends():
    """Admin: attendance trend data across all programs, grouped by date."""
    _ensure_enrollment_cache()
    start_str = request.args.get('start')
    end_str = request.args.get('end')
    today = date.today()

    # Defaults: current camp week, or last 7 days
    if not start_str or not end_str:
        cw = get_current_camp_week(today)
        if cw and cw in CAMP_WEEK_DATES:
            start_date = date.fromisoformat(CAMP_WEEK_DATES[cw][0])
            end_date = today
        else:
            start_date = today - timedelta(days=6)
            end_date = today
    else:
        try:
            start_date = date.fromisoformat(start_str)
            end_date = date.fromisoformat(end_str)
        except ValueError:
            start_date = today - timedelta(days=6)
            end_date = today

    # Query checkpoint_id=1 (Morning) only to avoid double-counting KC
    records = AttendanceRecord.query.filter(
        AttendanceRecord.checkpoint_id == 1,
        AttendanceRecord.date >= start_date,
        AttendanceRecord.date <= end_date
    ).all()

    # Group by date
    by_date = {}
    for r in records:
        d = r.date.isoformat()
        if d not in by_date:
            by_date[d] = {'present': 0, 'absent': 0, 'late': 0, 'early_pickup': 0}
        if r.status in by_date[d]:
            by_date[d][r.status] += 1

    # Calculate total enrolled per date (sum across all programs for that date's week)
    participants = {}
    if api_cache.get('data') and api_cache['data'].get('participants'):
        participants = api_cache['data']['participants']

    dates_result = []
    current = start_date
    while current <= end_date:
        # Skip weekends
        if current.weekday() < 5:
            d_str = current.isoformat()
            counts = by_date.get(d_str, {'present': 0, 'absent': 0, 'late': 0, 'early_pickup': 0})
            # Calculate total enrolled for this date's week
            wk = get_current_camp_week(current)
            total_enrolled = 0
            if wk and participants:
                wk_str = str(wk)
                for prog_name, weeks in participants.items():
                    if wk_str in weeks:
                        total_enrolled += len(weeks[wk_str])
            attended = counts['present'] + counts['late']
            rate = round(attended / total_enrolled * 100, 1) if total_enrolled > 0 else 0
            dates_result.append({
                'date': d_str,
                'present': counts['present'],
                'absent': counts['absent'],
                'late': counts['late'],
                'early_pickup': counts['early_pickup'],
                'total_enrolled': total_enrolled,
                'rate': rate
            })
        current += timedelta(days=1)

    return jsonify({
        'start': start_date.isoformat(),
        'end': end_date.isoformat(),
        'dates': dates_result
    })

@app.route('/api/attendance/detail/<program>')
@login_required
def attendance_detail(program):
    """Admin: individual camper attendance for a specific program on a date."""
    _ensure_enrollment_cache()
    target_date = _parse_date_param()

    # Get campers from enrollment
    week_num = get_current_camp_week(target_date)
    campers = []
    if api_cache.get('data') and api_cache['data'].get('participants'):
        participants = api_cache['data']['participants']
        if week_num:
            week_str = str(week_num)
            if program in participants and week_str in participants[program]:
                campers = participants[program][week_str]

    # Load person names from cache, auto-fetch missing ones from API
    all_pids = [c.get('personId') or c.get('person_id') for c in campers]
    persons_map = _load_and_fetch_persons(all_pids)

    # Trigger BAC sync in background if stale (never blocks this response)
    _ensure_bac_synced_background()

    # KC eligibility helper
    def _has_kc(person_id):
        entry = persons_map.get(str(person_id))
        if not isinstance(entry, dict):
            return False
        bac_weeks = entry.get('bac_weeks')
        if isinstance(bac_weeks, list) and week_num in bac_weeks:
            return True
        return False

    # Get records
    records = AttendanceRecord.query.filter_by(
        program_name=program, date=target_date
    ).all()
    att_map = {}
    for r in records:
        if r.person_id not in att_map:
            att_map[r.person_id] = {}
        att_map[r.person_id][str(r.checkpoint_id)] = {
            'status': r.status,
            'notes': r.notes,
            'recorded_by': r.recorded_by,
            'recorded_at': r.recorded_at.isoformat() if r.recorded_at else None
        }

    # Helper to extract name from persons_cache entry
    def _person_name(person_id, fallback=''):
        entry = persons_map.get(str(person_id))
        if isinstance(entry, dict):
            fn = entry.get('first_name', '')
            ln = entry.get('last_name', '')
            return f'{fn} {ln}'.strip() or fallback
        elif isinstance(entry, str):
            return entry or fallback
        return fallback

    camper_list = []
    for camper in campers:
        pid = str(camper.get('personId') or camper.get('person_id', ''))
        name = _person_name(pid, camper.get('name', f'Camper {pid}'))
        camper_list.append({
            'person_id': pid,
            'name': name,
            'has_kc': _has_kc(pid),
            'attendance': att_map.get(pid, {})
        })
    camper_list.sort(key=lambda c: c['name'].split()[-1] if c['name'] else '')

    checkpoints = _get_active_checkpoints()

    return jsonify({
        'program': program,
        'date': target_date.isoformat(),
        'week': week_num,
        'campers': camper_list,
        'checkpoints': [{'id': c.id, 'name': c.name, 'time_label': c.time_label} for c in checkpoints]
    })

# ---- Admin Management: Unit Leader Assignments ----

@app.route('/api/attendance/assignments', methods=['GET'])
@login_required
def get_assignments():
    """Admin: list all unit leader → program assignments."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    assignments = UnitLeaderAssignment.query.all()
    # Group by username
    by_user = {}
    for a in assignments:
        if a.username not in by_user:
            by_user[a.username] = []
        by_user[a.username].append(a.program_name)
    return jsonify({'assignments': by_user})

@app.route('/api/attendance/assignments', methods=['POST'])
@login_required
def save_assignment():
    """Admin: assign a program to a unit leader."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    username = data.get('username')
    program_name = data.get('program_name')
    if not username or not program_name:
        return jsonify({'error': 'Missing username or program_name'}), 400
    # Check user exists and is unit_leader
    user_acc = UserAccount.query.filter_by(username=username).first()
    if not user_acc:
        return jsonify({'error': 'User not found'}), 404
    # Check if already assigned
    existing = UnitLeaderAssignment.query.filter_by(username=username, program_name=program_name).first()
    if existing:
        return jsonify({'message': 'Already assigned'}), 200
    db.session.add(UnitLeaderAssignment(username=username, program_name=program_name))
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/attendance/assignments', methods=['DELETE'])
@login_required
def delete_assignment():
    """Admin: remove a program assignment from a unit leader."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    username = data.get('username')
    program_name = data.get('program_name')
    a = UnitLeaderAssignment.query.filter_by(username=username, program_name=program_name).first()
    if a:
        db.session.delete(a)
        db.session.commit()
    return jsonify({'success': True})

# ---- Admin Management: Checkpoints ----

@app.route('/api/attendance/checkpoints', methods=['PUT'])
@login_required
def update_checkpoint():
    """Admin: update a checkpoint's name, time_label, or active status."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    cp_id = data.get('id')
    cp = db.session.get(AttendanceCheckpoint, cp_id)
    if not cp:
        return jsonify({'error': 'Checkpoint not found'}), 404
    if 'name' in data:
        cp.name = data['name']
    if 'time_label' in data:
        cp.time_label = data['time_label']
    if 'active' in data:
        cp.active = bool(data['active'])
    if 'sort_order' in data:
        cp.sort_order = int(data['sort_order'])
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/attendance/week-info')
@login_required
def attendance_week_info():
    """Return current camp week info and all week date ranges."""
    today = date.today()
    current_week = get_current_camp_week(today)
    return jsonify({
        'today': today.isoformat(),
        'current_week': current_week,
        'is_camp_day': is_camp_day(today),
        'weeks': {str(k): {'start': v[0], 'end': v[1]} for k, v in CAMP_WEEK_DATES.items()}
    })

@app.route('/api/attendance/sync-bac', methods=['POST'])
@login_required
def sync_bac_data():
    """Sync Before and After Care data from CampMinder financial transactions + ECA sessions.
    Delegates to _sync_bac_to_cache() helper."""
    if current_user.role != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    if not CAMPMINDER_API_AVAILABLE:
        return jsonify({'error': 'CampMinder API not available'}), 500

    try:
        persons = _sync_bac_to_cache()
        kc_count = sum(1 for p in persons.values()
                       if isinstance(p, dict) and p.get('bac_weeks'))
        return jsonify({'success': True, 'total_kc_persons': kc_count})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': 'An internal error occurred'}), 500

# ==================== FIELD TRIP ROUTES ====================

@app.route('/fieldtrips/admin')
@login_required
def admin_fieldtrips():
    """Admin page for managing field trip venues, group-days, and assignments."""
    if not current_user.has_permission('manage_fieldtrips'):
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    return render_template('admin_fieldtrips.html',
                           user=current_user,
                           active_page='admin_fieldtrips')

def _get_fieldtrip_group_days():
    """Return the group-day mapping from GlobalSetting."""
    gs = GlobalSetting.query.filter_by(key='fieldtrip_group_days').first()
    if gs:
        try:
            return json.loads(gs.value)
        except Exception:
            pass
    return {
        'Monday': ['Teen Travel', 'Giborim', 'Madli-Teen'],
        'Tuesday': ['Teen Travel', 'Tnuah', 'Volleyball', 'Tiny Tnuah', 'Tsofim'],
        'Wednesday': ['Teen Travel', 'M&M', 'Tennis', 'Chaverim'],
        'Thursday': ['Teen Travel', 'Gymnastics', 'Art', 'Yeladim', 'Sports Academy', 'Karate'],
        'Friday': ['Teen Travel', 'Soccer', 'Basketball & Flag Football'],
    }

def _get_fieldtrip_kid_counts():
    """Compute kid counts per group per week from enrollment cache."""
    counts = {}  # {group_name: {week: count}}
    try:
        cache_path = os.path.join(DATA_FOLDER, 'api_cache.json')
        if not os.path.exists(cache_path):
            return counts
        with open(cache_path, 'r') as f:
            api_cache = json.load(f)
        participants = api_cache.get('data', {}).get('participants', [])
        if not participants:
            return counts

        # Build a mapping of fieldtrip group names to enrollment program names
        group_days = _get_fieldtrip_group_days()
        all_ft_groups = set()
        for day_groups in group_days.values():
            all_ft_groups.update(day_groups)

        # Get program settings for weeks_active
        prog_settings = {ps.program: ps for ps in ProgramSetting.query.all()}

        # For each participant, match to field trip groups by program name
        for p in participants:
            prog = p.get('programName', '')
            weeks = p.get('weeks', [])
            if not prog or not weeks:
                continue

            # Map enrollment program names to field trip group names
            # The field trip groups use shorter/different names
            ft_group = _map_program_to_ft_group(prog, all_ft_groups)
            if not ft_group:
                continue

            for w in weeks:
                wk = counts.setdefault(ft_group, {})
                wk[str(w)] = wk.get(str(w), 0) + 1
    except Exception:
        traceback.print_exc()
    return counts

def _map_program_to_ft_group(program_name, ft_groups):
    """Map an enrollment program name to a field trip group name."""
    # Direct match
    if program_name in ft_groups:
        return program_name

    # Common mappings from enrollment programs to field trip groups
    PROGRAM_TO_GROUP = {
        'Madli-Teen': 'Madli-Teen',
        "Children's Trust Madli-Teen": 'Madli-Teen',
        'Teen Travel': 'Teen Travel',
        'Teen Travel: Epic Trip to Orlando': 'Teen Travel',
        'Tsofim': 'Tsofim',
        "Children's Trust Tsofim": 'Tsofim',
        'Yeladim': 'Yeladim',
        "Children's Trust Yeladim": 'Yeladim',
        'Chaverim': 'Chaverim',
        "Children's Trust Chaverim": 'Chaverim',
        'Giborim': 'Giborim',
        "Children's Trust Giborim": 'Giborim',
        'Tnuah 1': 'Tnuah', 'Tnuah 2': 'Tnuah',
        'Extreme Tnuah': 'Tnuah',
        'Tiny Tnuah 1': 'Tiny Tnuah', 'Tiny Tnuah 2': 'Tiny Tnuah',
        'Teeny Tiny Tnuah': 'Tiny Tnuah',
        'Volleyball': 'Volleyball',
        'Tennis Academy': 'Tennis', 'Tennis Academy - Half Day': 'Tennis',
        'Tiny Tumblers Gymnastics': 'Gymnastics',
        'Recreational Gymnastics': 'Gymnastics',
        'Competitive Gymnastics Team': 'Gymnastics',
        'Art Exploration': 'Art',
        'Sports Academy 1': 'Sports Academy', 'Sports Academy 2': 'Sports Academy',
        'MMA Camp': 'Karate',
        'Soccer': 'Soccer',
        'Basketball': 'Basketball & Flag Football',
        'Flag Football': 'Basketball & Flag Football',
        'Music Camp': 'M&M',
        'Theater Camp': 'M&M',
    }
    return PROGRAM_TO_GROUP.get(program_name)


@app.route('/api/fieldtrips/matrix')
@login_required
def api_fieldtrips_matrix():
    """Return the full field trips matrix data."""
    if not current_user.has_permission('view_fieldtrips'):
        return jsonify({'error': 'Unauthorized'}), 403

    group_days = _get_fieldtrip_group_days()
    venues = {v.id: {'id': v.id, 'name': v.name, 'address': v.address,
                      'waiver_url': v.waiver_url}
              for v in FieldTripVenue.query.filter_by(active=True).all()}

    assignments = FieldTripAssignment.query.all()
    assignment_map = {}  # {group_name: {week_str: {...}}}
    for a in assignments:
        grp = assignment_map.setdefault(a.group_name, {})
        venue_info = venues.get(a.venue_id, {})
        grp[str(a.week)] = {
            'id': a.id,
            'venue_id': a.venue_id,
            'venue_name': venue_info.get('name', ''),
            'address': venue_info.get('address', ''),
            'waiver_url': venue_info.get('waiver_url', ''),
            'trip_date': a.trip_date.isoformat() if a.trip_date else None,
            'confirmed': a.confirmed,
            'comments': a.comments or '',
            'buses_ja': a.buses_ja or 0,
            'buses_jcc': a.buses_jcc or 0,
        }

    kid_counts = _get_fieldtrip_kid_counts()

    # Build ordered groups list by day
    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    ordered_groups = []
    for day in day_order:
        groups = group_days.get(day, [])
        for g in groups:
            ordered_groups.append({'name': g, 'day': day})

    return jsonify({
        'weeks': list(range(1, 10)),
        'week_dates': {str(k): v for k, v in CAMP_WEEK_DATES.items()},
        'day_order': day_order,
        'group_days': group_days,
        'groups': ordered_groups,
        'assignments': assignment_map,
        'kid_counts': kid_counts,
        'venues': list(venues.values()),
        'can_edit': current_user.has_permission('manage_fieldtrips'),
    })


@app.route('/api/fieldtrips/venues', methods=['GET'])
@login_required
def api_fieldtrips_venues_list():
    """List all venues."""
    if not current_user.has_permission('view_fieldtrips'):
        return jsonify({'error': 'Unauthorized'}), 403
    show_inactive = request.args.get('include_inactive', 'false') == 'true'
    query = FieldTripVenue.query
    if not show_inactive:
        query = query.filter_by(active=True)
    venues = query.order_by(FieldTripVenue.name).all()
    return jsonify({'venues': [
        {'id': v.id, 'name': v.name, 'address': v.address or '',
         'waiver_url': v.waiver_url or '', 'active': v.active}
        for v in venues
    ]})


@app.route('/api/fieldtrips/venues', methods=['POST'])
@login_required
def api_fieldtrips_venues_create():
    """Create a new venue."""
    if not current_user.has_permission('manage_fieldtrips'):
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Name is required'}), 400
    if FieldTripVenue.query.filter_by(name=name).first():
        return jsonify({'error': f'Venue "{name}" already exists'}), 400
    v = FieldTripVenue(
        name=name,
        address=(data.get('address') or '').strip(),
        waiver_url=(data.get('waiver_url') or '').strip() or None,
        active=True
    )
    db.session.add(v)
    db.session.commit()
    return jsonify({'success': True, 'venue': {'id': v.id, 'name': v.name,
                    'address': v.address or '', 'waiver_url': v.waiver_url or '', 'active': True}})


@app.route('/api/fieldtrips/venues/<int:venue_id>', methods=['PUT'])
@login_required
def api_fieldtrips_venues_update(venue_id):
    """Update a venue."""
    if not current_user.has_permission('manage_fieldtrips'):
        return jsonify({'error': 'Unauthorized'}), 403
    v = FieldTripVenue.query.get(venue_id)
    if not v:
        return jsonify({'error': 'Venue not found'}), 404
    data = request.get_json()
    if 'name' in data:
        new_name = (data['name'] or '').strip()
        if new_name and new_name != v.name:
            existing = FieldTripVenue.query.filter_by(name=new_name).first()
            if existing and existing.id != v.id:
                return jsonify({'error': f'Venue "{new_name}" already exists'}), 400
            v.name = new_name
    if 'address' in data:
        v.address = (data['address'] or '').strip()
    if 'waiver_url' in data:
        v.waiver_url = (data['waiver_url'] or '').strip() or None
    if 'active' in data:
        v.active = bool(data['active'])
    db.session.commit()
    return jsonify({'success': True, 'venue': {'id': v.id, 'name': v.name,
                    'address': v.address or '', 'waiver_url': v.waiver_url or '', 'active': v.active}})


@app.route('/api/fieldtrips/venues/<int:venue_id>', methods=['DELETE'])
@login_required
def api_fieldtrips_venues_delete(venue_id):
    """Soft-delete a venue (set active=False)."""
    if not current_user.has_permission('manage_fieldtrips'):
        return jsonify({'error': 'Unauthorized'}), 403
    v = FieldTripVenue.query.get(venue_id)
    if not v:
        return jsonify({'error': 'Venue not found'}), 404
    v.active = False
    db.session.commit()
    return jsonify({'success': True, 'message': f'Venue "{v.name}" deactivated'})


@app.route('/api/fieldtrips/assignments', methods=['PUT'])
@login_required
def api_fieldtrips_assignments_upsert():
    """Create or update a field trip assignment."""
    if not current_user.has_permission('manage_fieldtrips'):
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()
    group_name = (data.get('group_name') or '').strip()
    week = data.get('week')
    if not group_name or not week:
        return jsonify({'error': 'group_name and week are required'}), 400
    week = int(week)
    if week < 1 or week > 9:
        return jsonify({'error': 'Week must be 1-9'}), 400

    a = FieldTripAssignment.query.filter_by(group_name=group_name, week=week).first()
    if not a:
        a = FieldTripAssignment(group_name=group_name, week=week)
        db.session.add(a)

    venue_id = data.get('venue_id')
    if venue_id is not None:
        a.venue_id = int(venue_id) if venue_id else None
    if 'trip_date' in data:
        td = data['trip_date']
        a.trip_date = date.fromisoformat(td) if td else None
    if 'confirmed' in data:
        a.confirmed = bool(data['confirmed'])
    if 'comments' in data:
        a.comments = (data['comments'] or '').strip() or None
    if 'buses_ja' in data:
        a.buses_ja = int(data['buses_ja'] or 0)
    if 'buses_jcc' in data:
        a.buses_jcc = int(data['buses_jcc'] or 0)

    db.session.commit()
    # Return updated assignment
    venue = FieldTripVenue.query.get(a.venue_id) if a.venue_id else None
    return jsonify({'success': True, 'assignment': {
        'id': a.id,
        'group_name': a.group_name,
        'week': a.week,
        'venue_id': a.venue_id,
        'venue_name': venue.name if venue else '',
        'address': venue.address if venue else '',
        'waiver_url': venue.waiver_url if venue else '',
        'trip_date': a.trip_date.isoformat() if a.trip_date else None,
        'confirmed': a.confirmed,
        'comments': a.comments or '',
        'buses_ja': a.buses_ja or 0,
        'buses_jcc': a.buses_jcc or 0,
    }})


@app.route('/api/fieldtrips/group-days', methods=['GET'])
@login_required
def api_fieldtrips_group_days_get():
    """Get the group-day mapping."""
    if not current_user.has_permission('view_fieldtrips'):
        return jsonify({'error': 'Unauthorized'}), 403
    return jsonify({'group_days': _get_fieldtrip_group_days()})


@app.route('/api/fieldtrips/group-days', methods=['PUT'])
@login_required
def api_fieldtrips_group_days_update():
    """Update the group-day mapping."""
    if not current_user.has_permission('manage_fieldtrips'):
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.get_json()
    group_days = data.get('group_days')
    if not isinstance(group_days, dict):
        return jsonify({'error': 'group_days must be a dict'}), 400

    gs = GlobalSetting.query.filter_by(key='fieldtrip_group_days').first()
    if gs:
        gs.value = json.dumps(group_days)
    else:
        db.session.add(GlobalSetting(key='fieldtrip_group_days', value=json.dumps(group_days)))
    db.session.commit()
    return jsonify({'success': True, 'group_days': group_days})


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500

# ==================== MAIN ====================

if __name__ == '__main__':
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(DATA_FOLDER, exist_ok=True)
    
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('DEBUG', 'False') == 'True', threaded=True)
