"""
Budget Data Module for Camp Sol Taplin Finance Dashboard
FY2026 Budget + PO File Parser + Budget vs Actual Builder
"""

from io import BytesIO

# =============================================================================
# FY2026 HARDCODED BUDGET (from 2027.80.Camp V8 - FEB.xlsm "Dept. Totals")
# =============================================================================

BUDGET_FY2026 = {
    'revenue': {
        'total': 2_729_225,
        'program_fees': 2_680_850,
        'grants': 45_875,
        'contributions': 2_500,
        'discounts_scholarships': -248_163,  # budgeted discounts/scholarships
    },
    'expenses': {
        'total': 2_483_446,
        'categories': {
            'Salaries & Benefits': 1_690_566,
            'Food Service': 148_500,
            'Transportation': 81_800,
            'Field Trips': 68_600,
            'Program Supplies': 72_250,
            'Marketing': 25_500,
            'Staff Training': 12_000,
            'Security': 44_200,
            'Facility/Occupancy': 155_430,
            'Administrative': 51_000,
            'Specialists': 89_600,
            'Other Program': 44_000,
        }
    },
    'net': 245_779,
}

# =============================================================================
# PO CATEGORY → BUDGET CATEGORY MAPPING
# Maps each of the 35 PO categories to one of the 12 budget categories above
# =============================================================================

PO_TO_BUDGET_MAP = {
    'T-Shirts/Nike unit leader shirts': 'Program Supplies',
    'Food Lunches': 'Food Service',
    'Snacks/challah/staff seminars': 'Food Service',
    'Field Trip Admissions': 'Field Trips',
    'Transportation': 'Transportation',
    'Marketing': 'Marketing',
    'Staff training/Conferences': 'Staff Training',
    'Security JCC': 'Security',
    'Hillel Security': 'Security',
    'Backgroundscheck/Drug testing': 'Administrative',
    'Background checks/Drug testing': 'Administrative',
    'Backgroundscheck/drug testing': 'Administrative',
    'Backgrouns checks/Drug testing': 'Administrative',
    'Police': 'Security',
    'JCC housekeeping': 'Facility/Occupancy',
    'Hillel Housekeeping': 'Facility/Occupancy',
    'Friday events': 'Other Program',
    'Staff events/Awards': 'Other Program',
    'office supplies': 'Administrative',
    'Office supplies': 'Administrative',
    'Hillel damage': 'Facility/Occupancy',
    'Program supplies': 'Program Supplies',
    'Staff development': 'Staff Training',
    'Magic': 'Specialists',
    'Karate': 'Specialists',
    'I9 specialists (must submit W9)': 'Specialists',
    'Soccer specialist': 'Specialists',
    'Cleaning supplies/Hillel paper supplies': 'Facility/Occupancy',
    'LUNCH TENT AND FANS': 'Facility/Occupancy',
    "Children's trust supplies": 'Program Supplies',
    'Soccer camp FBS': 'Specialists',
    'Coding and Drone': 'Specialists',
    'Dance Specialists': 'Specialists',
    'Swim Right Staffing': 'Specialists',
    'Swim Right supplies': 'Program Supplies',
    'Birthday expenses': 'Other Program',
    'Campify': 'Other Program',
    'JDCN': 'Other Program',
    'Shlichim': 'Salaries & Benefits',
}

# Category display emojis
BUDGET_CATEGORY_EMOJIS = {
    'Salaries & Benefits': '👥',
    'Food Service': '🍽️',
    'Transportation': '🚌',
    'Field Trips': '🎢',
    'Program Supplies': '📦',
    'Marketing': '📣',
    'Staff Training': '🎓',
    'Security': '🔒',
    'Facility/Occupancy': '🏢',
    'Administrative': '📋',
    'Specialists': '⭐',
    'Other Program': '🎯',
}


# =============================================================================
# PO FILE PARSER
# =============================================================================

def parse_po_file(file_bytes):
    """
    Parse uploaded PO Excel file (2026 Camp PO.xlsx).
    Reads the 'PO Req form' sheet, extracts categories and spending.

    Returns list of dicts: [{category, balance, gl_code, sheet_num}, ...]
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    # Find the PO Req form sheet (try common names)
    sheet_name = None
    for name in wb.sheetnames:
        if 'po req' in name.lower() or 'req form' in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        # Fall back to second sheet (first is usually 'instructions')
        if len(wb.sheetnames) >= 2:
            sheet_name = wb.sheetnames[1]
        else:
            raise ValueError("Could not find 'PO Req form' sheet in the uploaded file")

    ws = wb[sheet_name]

    categories = []
    for row_idx in range(7, 50):  # rows 7-49 to cover all possible entries
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 10)]
        # row[0]=A (sheet#), row[1]=B (description), row[7]=H (GL), row[8]=I (balance)

        sheet_num = row[0]  # col A
        desc = row[1]       # col B - DESCRIPTION
        gl_code = row[7]    # col H - GL Coding
        balance = row[8]    # col I - BALANCE ON PO

        # Stop if we hit empty rows or summary rows
        if desc is None and sheet_num is None:
            continue
        if isinstance(desc, str) and desc.strip().upper() in ('BUDGET TOTALS', 'APPROVED TO DATE', 'SPENT TO DATE', 'DEPT APPROVED BALANCE', ''):
            continue

        if desc:
            # Balance is typically negative (money spent), we want positive "spent" amount
            spent = 0
            if isinstance(balance, (int, float)):
                spent = abs(float(balance))
            elif isinstance(balance, str):
                try:
                    cleaned = balance.replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip()
                    if cleaned and cleaned != '-':
                        spent = abs(float(cleaned))
                except (ValueError, TypeError):
                    spent = 0

            categories.append({
                'sheet_num': sheet_num,
                'category': str(desc).strip(),
                'balance': round(spent, 2),
                'gl_code': str(gl_code).strip() if gl_code else '',
            })

    wb.close()
    return categories


# =============================================================================
# BUDGET VS ACTUAL BUILDER
# =============================================================================

def build_budget_vs_actual(po_categories):
    """
    Group PO spending by budget category and compare with budgeted amounts.

    Returns:
    {
        'categories': [{category, emoji, budgeted, actual, variance, pct_used, po_items}, ...],
        'totals': {budgeted, actual, variance, pct_used},
        'po_detail': [{category, balance, budget_group}, ...]
    }
    """
    # Aggregate actuals by budget category
    actual_by_budget = {}
    po_items_by_budget = {}

    for po in po_categories:
        # Try exact match first, then case-insensitive
        budget_cat = PO_TO_BUDGET_MAP.get(po['category'])
        if not budget_cat:
            # Try case-insensitive match
            for key, val in PO_TO_BUDGET_MAP.items():
                if key.lower() == po['category'].lower():
                    budget_cat = val
                    break
        if not budget_cat:
            budget_cat = 'Other Program'

        actual_by_budget.setdefault(budget_cat, 0)
        actual_by_budget[budget_cat] += po['balance']

        po_items_by_budget.setdefault(budget_cat, [])
        po_items_by_budget[budget_cat].append(po['category'])

    # Build comparison rows
    result = []
    for cat, budgeted in BUDGET_FY2026['expenses']['categories'].items():
        actual = round(actual_by_budget.get(cat, 0), 2)
        variance = round(budgeted - actual, 2)
        pct_used = round((actual / budgeted * 100), 1) if budgeted > 0 else 0

        result.append({
            'category': cat,
            'emoji': BUDGET_CATEGORY_EMOJIS.get(cat, '📌'),
            'budgeted': budgeted,
            'actual': actual,
            'variance': variance,
            'pct_used': pct_used,
            'po_items': po_items_by_budget.get(cat, []),
        })

    # Sort by actual spending (highest first)
    result.sort(key=lambda x: x['actual'], reverse=True)

    # Totals
    total_actual = round(sum(r['actual'] for r in result), 2)
    total_budgeted = BUDGET_FY2026['expenses']['total']
    totals = {
        'budgeted': total_budgeted,
        'actual': total_actual,
        'variance': round(total_budgeted - total_actual, 2),
        'pct_used': round(total_actual / total_budgeted * 100, 1) if total_budgeted > 0 else 0,
    }

    # PO detail with budget group assignment
    po_detail = []
    for po in po_categories:
        budget_cat = PO_TO_BUDGET_MAP.get(po['category'], 'Other Program')
        po_detail.append({
            'category': po['category'],
            'balance': po['balance'],
            'gl_code': po['gl_code'],
            'budget_group': budget_cat,
        })

    return {
        'categories': result,
        'totals': totals,
        'po_detail': po_detail,
    }
